"""Verify Kamlesh's row in all 3 generated Excel files."""
import openpyxl

print("=" * 90)
print("VERIFICATION: Kamlesh (EMP-021) — July 2026")
print("=" * 90)

EXPECTED = {
    'present': 24,
    'absent': 3,
    'sundays': 4,
    'total_days': 31,
    'worked_hrs_incl_ot': '216:04',
    'sunday_hrs': '36:00',
    'total_hrs': '252:04',
    'gross_salary_approx': 17166,
}

# ── File 1: Payroll_Summary_July_2026.xlsx ──
print("\n[1] Payroll_Summary_July_2026.xlsx → 'Payroll Register' sheet")
wb = openpyxl.load_workbook('/home/z/my-project/download/Payroll_Summary_July_2026.xlsx', data_only=False)
ws = wb['Payroll Register']
print(f"    Header R4: {[ws.cell(4, c).value for c in range(1, 15)]}")
for r in range(5, ws.max_row + 1):
    name = ws.cell(r, 2).value
    if 'Kamlesh' in str(name):
        print(f"    Row {r}:")
        for c, label in [(3,'Monthly Salary'),(4,'Working Hrs'),(5,'Sl/Hr'),(6,'Present'),(7,'Absent'),(8,'Worked Hrs incl OT'),(9,'Additional hrs (Sunday)'),(10,'Total Hrs'),(11,'Gross'),(14,'Net')]:
            print(f"      {label:25s}: {ws.cell(r, c).value}")
        # Verify
        p = ws.cell(r, 6).value
        a = ws.cell(r, 7).value
        print(f"\n    ✓ Present = {p} (expected {EXPECTED['present']}) → {'PASS' if p==EXPECTED['present'] else 'FAIL'}")
        print(f"    ✓ Absent  = {a} (expected {EXPECTED['absent']})   → {'PASS' if a==EXPECTED['absent'] else 'FAIL'}")
        # Days sum check
        # Total days = Present + Absent + Sundays
        print(f"    ✓ Total days = {p}+{a}+{EXPECTED['sundays']} = {p+a+EXPECTED['sundays']} (expected {EXPECTED['total_days']})")
        break

# ── File 2: Payroll_Master_July_2026.xlsx ──
print("\n[2] Payroll_Master_July_2026.xlsx → 'Master' sheet")
wb = openpyxl.load_workbook('/home/z/my-project/download/Payroll_Master_July_2026.xlsx', data_only=False)
print(f"    Sheets: {wb.sheetnames}")
ws = wb['Master']
# print header
for r in [1,2,3,4]:
    row_vals = [ws.cell(r, c).value for c in range(1, 16)]
    if any(v is not None for v in row_vals):
        print(f"    R{r}: {row_vals}")
for r in range(5, ws.max_row + 1):
    name = ws.cell(r, 3).value
    if 'Kamlesh' in str(name):
        print(f"\n    Kamlesh row {r}:")
        for c in range(1, ws.max_column + 1):
            hdr = ws.cell(4, c).value or ws.cell(3, c).value or f"C{c}"
            v = ws.cell(r, c).value
            print(f"      {hdr:25s}: {v}")
        break

# ── File 3: Attendance_Tracker_Monthly_July_2026.xlsx ──
print("\n[3] Attendance_Tracker_Monthly_July_2026.xlsx → 'LAPL' sheet")
wb = openpyxl.load_workbook('/home/z/my-project/download/Attendance_Tracker_Monthly_July_2026.xlsx', data_only=False)
print(f"    Sheets: {wb.sheetnames}")
ws = wb['LAPL']
for r in [1,2,3,4,5]:
    row_vals = [ws.cell(r, c).value for c in range(1, 12)]
    if any(v is not None for v in row_vals):
        print(f"    R{r}: {row_vals}")
# Find Kamlesh
for r in range(5, ws.max_row + 1):
    name = ws.cell(r, 1).value or ws.cell(r, 2).value
    if 'Kamlesh' in str(name):
        print(f"\n    Kamlesh row {r}:")
        # Print all non-empty cells with their column header
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and v != '':
                h1 = ws.cell(1, c).value
                h2 = ws.cell(2, c).value
                h3 = ws.cell(3, c).value
                h4 = ws.cell(4, c).value
                print(f"      Col {c} (hdrs: {h1} | {h2} | {h3} | {h4}): {v}")
        break

print("\n" + "=" * 90)
print("DONE")
