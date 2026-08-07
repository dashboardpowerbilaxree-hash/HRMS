#!/usr/bin/env python3
"""Generate two files:
1. Payroll_Master_July_2026.xlsx — Master + per-firm salary sheets (LAPL/LRSL/SDF/SI _July_2026_Sal)
2. Attendance_Tracker_Monthly_July_2026.xlsx — per-firm daily attendance grids (P/A/Half/OT codes)
"""
import json
import os
import calendar
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2

DATA = json.load(open('/home/z/my-project/scripts/july_data.json'))
META = DATA['meta']
EMPLOYEES = DATA['employees']
EMPLOYEES.sort(key=lambda e: (e['firm'] or '~', e['fullName']))

YEAR = META['year']
MONTH = META['month']
DAYS = META['daysInMonth']
SUNDAYS = set(META['sundays'])
HOLIDAYS = set(META['holidays'])

# Reconnect to get raw attendance for daily grids
DB_URL = open('/tmp/db_url.txt').read().strip()
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()
cur.execute("""
    SELECT "employeeId", date, "totalHours", "overtimeHours", status, "halfDay"
    FROM "Attendance"
    WHERE date >= '2026-07-01' AND date < '2026-08-01'
""")
ATT = {}
for r in cur.fetchall():
    emp_id = r[0]
    day = r[1].day
    ATT.setdefault(emp_id, {})[day] = {
        'totalHours': float(r[2] or 0),
        'overtimeHours': float(r[3] or 0),
        'status': r[4],
        'halfDay': r[5],
    }
conn.close()

# ──────────── STYLES ────────────
NAVY = "1F3A5F"
GOLD = "C9A961"
LIGHT_BLUE = "E8EEF4"
LIGHT_GOLD = "FAF5E8"
LIGHT_GREEN = "E8F5E9"
LIGHT_RED = "FDEDEC"
LIGHT_YELLOW = "FEF9E7"
SUN_FILL = "F4E5D6"   # warm tan for Sunday columns
WHITE = "FFFFFF"

thin = Side(border_style="thin", color="666666")
medium = Side(border_style="medium", color=NAVY)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

font_title = Font(name='Calibri', size=14, bold=True, color=WHITE)
font_subtitle = Font(name='Calibri', size=10, italic=True, color=WHITE)
font_header = Font(name='Calibri', size=9, bold=True, color=WHITE)
font_body = Font(name='Calibri', size=10)
font_body_sm = Font(name='Calibri', size=9)
font_total = Font(name='Calibri', size=10, bold=True, color=NAVY)

fill_title = PatternFill("solid", fgColor=NAVY)
fill_header = PatternFill("solid", fgColor=NAVY)
fill_total = PatternFill("solid", fgColor=LIGHT_GOLD)
fill_alt = PatternFill("solid", fgColor=LIGHT_BLUE)
fill_sun = PatternFill("solid", fgColor=SUN_FILL)
fill_present = PatternFill("solid", fgColor=LIGHT_GREEN)
fill_absent = PatternFill("solid", fgColor=LIGHT_RED)
fill_half = PatternFill("solid", fgColor=LIGHT_YELLOW)

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

def fmt_hrs(v):
    if v is None or v == 0:
        return "0:00"
    h = int(v)
    m = int(round((v - h) * 60))
    if m == 60:
        h += 1
        m = 0
    return f"{h}:{m:02d}"

def status_code(day, rec):
    """Return short code for daily grid."""
    if day in SUNDAYS:
        return "S"  # Sunday
    if not rec:
        return "A"  # Absent (no record)
    s = rec['status']
    if s == 'absent':
        return "A"
    if s in ('half-day', 'half_day') or rec.get('halfDay'):
        return "H"  # Half-day
    if s == 'late':
        return "L"  # Late
    if s == 'early-out':
        return "E"  # Early out
    if s == 'present':
        return "P"  # Present
    return s[:1].upper() if s else "-"

# ════════════════════════════════════════════════════════════════════════════
# FILE 1: Payroll_Master_July_2026.xlsx
# ════════════════════════════════════════════════════════════════════════════
wb1 = Workbook()
wb1.properties.creator = "Z.ai"

# ─── Master Sheet ───
ws = wb1.active
ws.title = "Master"

ws.merge_cells('A1:L1')
ws['A1'] = "LAXREE GROUP OF COMPANIES — Master Employee List"
ws['A1'].font = font_title
ws['A1'].fill = fill_title
ws['A1'].alignment = align_center
ws.row_dimensions[1].height = 28

ws.merge_cells('A2:L2')
ws['A2'] = f"Salary Month: July 2026  |  Days in Month: {DAYS}  |  Sundays: {len(SUNDAYS)}  |  Holidays: {len(HOLIDAYS)}  |  Working Days: {META['workingDays']}"
ws['A2'].font = Font(name='Calibri', size=10, italic=True, color=WHITE)
ws['A2'].fill = fill_title
ws['A2'].alignment = align_center

master_headers = ['Emp Code', 'Full Name', 'Firm', 'Location', 'Salary Type',
                  'Monthly Salary (₹)', 'Daily Rate (₹)', 'Shift Hrs/Day',
                  'Shift Start', 'Shift End', 'Employment Type', 'Status']
for i, h in enumerate(master_headers, 1):
    c = ws.cell(4, i, h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center
    c.border = border_all
ws.row_dimensions[4].height = 32

r = 5
for emp in EMPLOYEES:
    vals = [
        emp['employeeId'],
        emp['fullName'],
        emp['firm'] or '-',
        emp['location'] or '-',
        emp['salaryType'] or 'Monthly',
        emp['monthlySalary'],
        round(emp['monthlySalary'] / DAYS, 2) if emp['monthlySalary'] else 0,
        emp['shiftHours'],
        emp['shiftStart'],
        emp['shiftEnd'],
        emp['employmentType'] or 'Full Time',
        'Active',
    ]
    for i, v in enumerate(vals, 1):
        c = ws.cell(r, i, v)
        c.font = font_body
        c.border = border_all
        if i in (1, 3, 7, 8, 9, 10, 11, 12):
            c.alignment = align_center
        elif i in (6, 7):
            c.alignment = align_right
            c.number_format = '#,##0.00'
        else:
            c.alignment = align_left
        if r % 2 == 0:
            c.fill = fill_alt
    r += 1

# Total
ws.cell(r, 2).value = "TOTAL"
ws.cell(r, 6).value = f"=SUM(F5:F{r-1})"
ws.cell(r, 6).number_format = '#,##0'
for i in range(1, 13):
    c = ws.cell(r, i)
    c.font = font_total
    c.fill = fill_total
    c.border = border_all
    if i in (6, 7):
        c.alignment = align_right
    else:
        c.alignment = align_center

widths = [10, 24, 7, 16, 11, 16, 13, 11, 11, 11, 14, 9]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ─── Per-firm Salary Sheets ───
firms = sorted(set(e['firm'] for e in EMPLOYEES if e['firm']))
print(f"Firms: {firms}")

for firm in firms:
    firm_emps = [e for e in EMPLOYEES if e['firm'] == firm]
    sheet_name = f"{firm}_July_2026_Sal"
    ws = wb1.create_sheet(sheet_name)

    ws.merge_cells('A1:X1')
    ws['A1'] = f"{firm} — Salary Sheet: July 2026"
    ws['A1'].font = font_title
    ws['A1'].fill = fill_title
    ws['A1'].alignment = align_center
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:X2')
    ws['A2'] = (f"Salary Month: July 2026  |  Days in Month: {DAYS}  |  "
                f"Sundays: {len(SUNDAYS)}  |  Holidays: {len(HOLIDAYS)}  |  "
                f"Working Days: {META['workingDays']}  |  Employees: {len(firm_emps)}")
    ws['A2'].font = Font(name='Calibri', size=9, italic=True, color="555555")
    ws['A2'].alignment = align_center

    sal_headers = [
        'S.No', 'Emp Code', 'Employee Name', 'Location', 'Monthly Salary (₹)',
        'Shift Hrs/Day', 'Sl/Hr (₹)', 'Worked Hrs including OT', 'OT Hours',
        'OT Rate/Hr (₹)', 'OT Amount (₹)', 'Additional Hrs (Sundays)', 'PH Hours',
        'Total Hrs', 'Gross Salary (₹)', 'TDS (₹)', 'Loan/Amt (₹)', 'Advance (₹)',
        'Security Deposit (₹)', 'Total Deductions (₹)', 'Balance Loan Amt (₹)',
        'Arrear (₹)', 'Net Salary (₹)', 'Remarks'
    ]
    for i, h in enumerate(sal_headers, 1):
        c = ws.cell(4, i, h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_all
    ws.row_dimensions[4].height = 42

    r = 5
    for idx, emp in enumerate(firm_emps, 1):
        # All values: H, I, L, M as numeric hours; J=G (OT rate = Sl/Hr); K=I*J; N=H+L+M (WorkedHrs incl OT + Sunday + PH); O=N*G; T=SUM(P:S); W=O-T+V
        # NOTE: Total Hrs = H+L+M (NOT +I). H already includes OT, so adding I separately would double-count OT.
        vals = [
            idx,                                                        # A: S.No
            emp['employeeId'],                                          # B: Emp Code
            emp['fullName'],                                            # C: Name
            emp['location'] or '-',                                     # D: Location
            emp['monthlySalary'],                                       # E: Monthly Salary
            emp['shiftHours'],                                          # F: Shift Hrs/Day
            None,                                                       # G: Sl/Hr — formula
            round(emp['workedHrsInclOT'], 2),                           # H: Worked Hrs incl OT
            round(emp['otHrs'], 2),                                     # I: OT Hours
            None,                                                       # J: OT Rate — formula =G
            None,                                                       # K: OT Amount — formula =I*J
            emp['sundayHrs'],                                           # L: Additional Hrs (Sundays)
            0,                                                          # M: PH Hours
            None,                                                       # N: Total Hrs — formula
            None,                                                       # O: Gross — formula
            0,                                                          # P: TDS
            0,                                                          # Q: Loan
            0,                                                          # R: Advance
            0,                                                          # S: Security Deposit
            None,                                                       # T: Total Deductions — formula
            0,                                                          # U: Balance Loan Amt
            0,                                                          # V: Arrear
            None,                                                       # W: Net Salary — formula
            '',                                                         # X: Remarks
        ]
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v)
            c.font = font_body_sm
            c.border = border_all
            if i in (1, 2, 4, 6, 9, 12, 13):
                c.alignment = align_center
            elif i in (5, 7, 8, 10, 11, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23):
                c.alignment = align_right
            else:
                c.alignment = align_left
            if r % 2 == 0:
                c.fill = fill_alt

        # Formulas
        ws.cell(r, 7).value = f"=E{r}/({DAYS}*F{r})"          # G: Sl/Hr
        ws.cell(r, 7).number_format = '0.00'
        ws.cell(r, 10).value = f"=G{r}"                       # J: OT Rate = Sl/Hr
        ws.cell(r, 10).number_format = '0.00'
        ws.cell(r, 11).value = f"=I{r}*J{r}"                  # K: OT Amount
        ws.cell(r, 11).number_format = '0.00'
        ws.cell(r, 14).value = f"=H{r}+L{r}+M{r}"          # N: Total Hrs (WorkedHrs incl OT + Sunday + PH — NO separate OT, H already has it)
        ws.cell(r, 14).number_format = '0.00'
        ws.cell(r, 15).value = f"=N{r}*G{r}"                  # O: Gross Salary
        ws.cell(r, 15).number_format = '#,##0.00'
        ws.cell(r, 20).value = f"=SUM(P{r}:S{r})"             # T: Total Deductions
        ws.cell(r, 20).number_format = '0.00'
        ws.cell(r, 23).value = f"=O{r}-T{r}+IF(V{r}=\"\",0,V{r})"  # W: Net Salary
        ws.cell(r, 23).number_format = '#,##0.00'

        ws.cell(r, 5).number_format = '#,##0'
        ws.cell(r, 8).number_format = '0.00'
        ws.cell(r, 9).number_format = '0.00'
        ws.cell(r, 12).number_format = '0.00'
        ws.cell(r, 13).number_format = '0.00'
        r += 1

    # Total row
    ws.cell(r, 3).value = "TOTAL"
    ws.cell(r, 5).value = f"=SUM(E5:E{r-1})"
    ws.cell(r, 8).value = f"=SUM(H5:H{r-1})"
    ws.cell(r, 9).value = f"=SUM(I5:I{r-1})"
    ws.cell(r, 11).value = f"=SUM(K5:K{r-1})"
    ws.cell(r, 12).value = f"=SUM(L5:L{r-1})"
    ws.cell(r, 14).value = f"=SUM(N5:N{r-1})"
    ws.cell(r, 15).value = f"=SUM(O5:O{r-1})"
    ws.cell(r, 20).value = f"=SUM(T5:T{r-1})"
    ws.cell(r, 23).value = f"=SUM(W5:W{r-1})"
    for i in range(1, 25):
        c = ws.cell(r, i)
        c.font = font_total
        c.fill = fill_total
        c.border = border_all
        if i in (5, 8, 9, 11, 12, 14, 15, 20, 23):
            c.alignment = align_right
        else:
            c.alignment = align_center
    ws.cell(r, 5).number_format = '#,##0'
    for col in (8, 9, 11, 12, 14, 15, 20, 23):
        ws.cell(r, col).number_format = '#,##0.00'

    widths = [5, 10, 22, 14, 14, 10, 9, 12, 9, 10, 11, 13, 9, 10, 14, 9, 10, 10, 13, 13, 13, 11, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E5"

OUT1 = '/home/z/my-project/download/Payroll_Master_July_2026.xlsx'
wb1.save(OUT1)
print(f"✓ Saved: {OUT1}")

# ════════════════════════════════════════════════════════════════════════════
# FILE 2: Attendance_Tracker_Monthly_July_2026.xlsx
# ════════════════════════════════════════════════════════════════════════════
wb2 = Workbook()
wb2.properties.creator = "Z.ai"

for firm in firms:
    firm_emps = [e for e in EMPLOYEES if e['firm'] == firm]
    sheet_name = f"{firm}_July_2026_Att"
    ws = wb2.create_sheet(sheet_name) if firm != firms[0] or len(wb2.sheetnames) > 1 else wb2.active
    if ws is wb2.active and ws.title == "Sheet":
        ws.title = sheet_name

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=DAYS+8)
    ws.cell(1, 1).value = f"LAXREE GROUP OF COMPANIES — {firm} Attendance Tracker — July 2026"
    ws.cell(1, 1).font = font_title
    ws.cell(1, 1).fill = fill_title
    ws.cell(1, 1).alignment = align_center
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=DAYS+8)
    ws.cell(2, 1).value = (f"Days in Month: {DAYS}  |  Sundays: {len(SUNDAYS)}  |  "
                           f"Holidays: {len(HOLIDAYS)}  |  Working Days: {META['workingDays']}  |  "
                           f"Codes: P=Present  A=Absent  H=Half-day  L=Late  E=Early-out  S=Sunday")
    ws.cell(2, 1).font = Font(name='Calibri', size=9, italic=True, color="555555")
    ws.cell(2, 1).alignment = align_center

    # Header row 3 + 4 (day numbers + weekday names)
    fixed_headers = ['S.No', 'Emp Code', 'Employee Name']
    for i, h in enumerate(fixed_headers, 1):
        c = ws.cell(3, i, h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_all
        ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)

    # Day columns
    for d in range(1, DAYS+1):
        col = 3 + d
        c = ws.cell(3, col, d)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_all
        weekday = calendar.weekday(YEAR, MONTH, d)
        day_name = calendar.day_abbr[weekday][:2]
        c2 = ws.cell(4, col, day_name)
        c2.font = Font(name='Calibri', size=8, bold=True, color=WHITE)
        c2.fill = fill_header
        c2.alignment = align_center
        c2.border = border_all

    # Summary columns
    summary_start = 3 + DAYS + 1
    summary_headers = ['Present', 'Absent', 'Half', 'Total Hrs', 'OT Hrs']
    for i, h in enumerate(summary_headers):
        col = summary_start + i
        c = ws.cell(3, col, h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_all
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 16

    # Data rows
    r = 5
    for idx, emp in enumerate(firm_emps, 1):
        ws.cell(r, 1).value = idx
        ws.cell(r, 2).value = emp['employeeId']
        ws.cell(r, 3).value = emp['fullName']
        for c in range(1, 4):
            cell = ws.cell(r, c)
            cell.font = font_body_sm
            cell.border = border_all
            cell.alignment = align_center if c != 3 else align_left
            if r % 2 == 0:
                cell.fill = fill_alt

        emp_att = ATT.get(emp['employeeId'], {})
        for d in range(1, DAYS+1):
            col = 3 + d
            rec = emp_att.get(d)
            code = status_code(d, rec)
            c = ws.cell(r, col, code)
            c.font = font_body_sm
            c.border = border_all
            c.alignment = align_center
            if d in SUNDAYS:
                c.fill = fill_sun
            elif code == 'P':
                c.fill = fill_present
            elif code == 'A':
                c.fill = fill_absent
            elif code == 'H':
                c.fill = fill_half

        # Summary
        ws.cell(r, summary_start).value = emp['presentDays']
        ws.cell(r, summary_start+1).value = emp['absentDays']
        ws.cell(r, summary_start+2).value = emp['halfDays']
        ws.cell(r, summary_start+3).value = fmt_hrs(emp['workedHrsInclOT'])
        ws.cell(r, summary_start+4).value = fmt_hrs(emp['otHrs'])
        for i in range(5):
            cell = ws.cell(r, summary_start+i)
            cell.font = font_total if i < 3 else font_body_sm
            cell.border = border_all
            cell.alignment = align_center
            if r % 2 == 0:
                cell.fill = fill_alt
        r += 1

    # Total row
    ws.cell(r, 3).value = "TOTAL"
    for i in range(3):
        col = summary_start + i
        if i == 0:
            ws.cell(r, col).value = f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{r-1})"
        elif i == 1:
            ws.cell(r, col).value = f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{r-1})"
        elif i == 2:
            ws.cell(r, col).value = f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{r-1})"
    for i in range(1, DAYS + 8 + 1):
        c = ws.cell(r, i)
        c.font = font_total
        c.fill = fill_total
        c.border = border_all
        c.alignment = align_center

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 22
    for d in range(1, DAYS+1):
        ws.column_dimensions[get_column_letter(3+d)].width = 4
    for i in range(5):
        ws.column_dimensions[get_column_letter(summary_start+i)].width = 9

    ws.freeze_panes = "D5"

OUT2 = '/home/z/my-project/download/Attendance_Tracker_Monthly_July_2026.xlsx'
wb2.save(OUT2)
print(f"✓ Saved: {OUT2}")

# ─── Kamlesh verification ───
kam = next(e for e in EMPLOYEES if 'Kamlesh' in e['fullName'])
print(f"\n{'='*80}")
print(f"KAMLESH VERIFICATION:")
print(f"{'='*80}")
print(f"  Present: {kam['presentDays']}  |  Absent: {kam['absentDays']}  |  Half: {kam['halfDays']}")
print(f"  Worked Hrs incl OT: {fmt_hrs(kam['workedHrsInclOT'])}")
print(f"  Sunday Hrs: {fmt_hrs(kam['sundayHrs'])}")
print(f"  Total Hrs: {fmt_hrs(kam['totalHrs'])}")
print(f"  Gross Salary: ₹{kam['grossSalary']:,.2f}")
print(f"  Absent Days (per grid): {[d for d in range(1, DAYS+1) if status_code(d, ATT.get(kam['employeeId'], {}).get(d)) == 'A' and d not in SUNDAYS]}")
