#!/usr/bin/env python3
"""Verify that the Leave column in the newly generated Payroll_Master_July_2026.xlsx
matches the production export-master at /tmp/prod-master.xlsx for all 42 employees.
"""
from openpyxl import load_workbook

FIRM_COUNTS = {'LAPL': 15, 'LRSL': 16, 'SDF': 5, 'SI': 6}


def find_section3_leave(ws, num_emps):
    """Find the section 3 header row by locating 'Total Working Hours' (which
    only appears in section 3 header). Leave column is the next col.
    Then return {name_lower: leave_value} for each employee row.
    """
    twh_row = None
    twh_col = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v == 'Total Working Hours':
                twh_row = r
                twh_col = c
                break
        if twh_row:
            break
    if twh_row is None:
        return {}

    # Leave column is the next column after Total Working Hours
    leave_col = twh_col + 1

    # Employee rows are twh_row + 2 ... twh_row + 2 + num_emps - 1
    # (skip the header row itself + the IN/OUT/TOTAL HRS sub-header row)
    result = {}
    for r in range(twh_row + 2, twh_row + 2 + num_emps):
        name = ws.cell(row=r, column=1).value
        if name:
            lv = ws.cell(row=r, column=leave_col).value
            result[str(name).lower().strip()] = lv
    return result


# Load both Excel files
wb_new = load_workbook('/home/z/my-project/download/Payroll_Master_July_2026.xlsx', data_only=False)
wb_prod = load_workbook('/tmp/prod-master.xlsx', data_only=False)

print(f'{"Firm":5} {"Employee":25} {"New Leave":10} {"Prod Leave":12} {"Match":6}')
print('-' * 70)

total_mismatches = 0
total_checked = 0
for firm in ['LAPL', 'LRSL', 'SDF', 'SI']:
    num = FIRM_COUNTS[firm]
    new_sheet = f'{firm}_July_2026_Sal'
    prod_sheet = firm

    ws_new = wb_new[new_sheet]
    ws_prod = wb_prod[prod_sheet]

    new_lv = find_section3_leave(ws_new, num)
    prod_lv = find_section3_leave(ws_prod, num)

    all_names = set(new_lv.keys()) | set(prod_lv.keys())
    for name in sorted(all_names):
        n = new_lv.get(name)
        p = prod_lv.get(name)
        match = 'OK'
        try:
            if n is None or p is None:
                match = 'MISSING'
                total_mismatches += 1
            elif abs(float(n) - float(p)) > 0.01:
                match = 'MISMATCH'
                total_mismatches += 1
        except (ValueError, TypeError):
            match = 'PARSE_ERR'
            total_mismatches += 1

        print(f'{firm:5} {name[:25]:25} {str(n):10} {str(p):12} {match}')
        total_checked += 1

print('-' * 70)
print(f'Total checked: {total_checked}, Mismatches: {total_mismatches}')
if total_mismatches == 0:
    print('✅ ALL 42 employees: Leave column matches production export-master!')
else:
    print('❌ Some mismatches found — review above.')
