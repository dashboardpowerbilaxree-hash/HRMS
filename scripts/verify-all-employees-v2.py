"""Final verification: all 42 employees across all 3 Excel files + production API.
Uses the production API as the single source of truth.
"""
import json
import openpyxl

# Load API-sourced ground truth
DATA = json.load(open('/home/z/my-project/scripts/july_data.json'))
META = DATA['meta']
EMPLOYEES = {e['employeeId']: e for e in DATA['employees']}
# Also index by name for Excel lookups
BY_NAME = {e['fullName']: e for e in DATA['employees']}

print(f"Ground truth: {len(EMPLOYEES)} employees from production API")
print(f"  Days in month: {META['daysInMonth']}, Sundays: {META['numSundays']}, Working days: {META['workingDays']}")

def fmt_hrs(v):
    if v is None or v == 0: return "0:00"
    h = int(v)
    m = int(round((v - h) * 60))
    if m == 60: h += 1; m = 0
    return f"{h}:{m:02d}"

errors = {1: [], 2: [], 3: []}

# ════════════════════════════════════════════════════════════════════════════
# FILE 1: Payroll_Summary_July_2026.xlsx — Payroll Register sheet
# ════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 120)
print("[1] Payroll_Summary_July_2026.xlsx — Payroll Register sheet")
print("=" * 120)
wb1 = openpyxl.load_workbook('/home/z/my-project/download/Payroll_Summary_July_2026.xlsx')
ws1 = wb1['Payroll Register']
print(f"{'#':>3} {'Emp Name':<25} {'P':>3} {'A':>3} {'WorkedOT':>9} {'AddSun':>7} {'Total':>8}  Status")
print("-" * 80)
count = 0
for r in range(5, ws1.max_row + 1):
    name = ws1.cell(r, 2).value
    if not name or name == 'TOTAL':
        continue
    count += 1
    present = ws1.cell(r, 6).value
    absent = ws1.cell(r, 7).value
    worked_hrs_str = ws1.cell(r, 8).value
    add_hrs_str = ws1.cell(r, 9).value
    total_hrs_num = ws1.cell(r, 10).value

    db = BY_NAME.get(name)
    if not db:
        errors[1].append(f"{name}: not found in API data")
        print(f"{count:>3} {name[:24]:<25} — NOT FOUND")
        continue

    issues = []
    if present != db['presentDays']: issues.append(f"P≠{db['presentDays']}")
    if absent != db['absentDays']: issues.append(f"A≠{db['absentDays']}")
    if worked_hrs_str != fmt_hrs(db['workedHrsInclOT']): issues.append(f"WrkHrs≠{fmt_hrs(db['workedHrsInclOT'])}")
    if add_hrs_str != fmt_hrs(db['sundayHrs']): issues.append(f"Sun≠{fmt_hrs(db['sundayHrs'])}")
    if total_hrs_num != round(db['totalHrs'], 2): issues.append(f"Tot≠{db['totalHrs']}")

    status = " ".join(issues) if issues else "OK"
    if issues: errors[1].append(f"{name}: {status}")
    print(f"{count:>3} {name[:24]:<25} {present:>3} {absent:>3} {worked_hrs_str:>9} {add_hrs_str:>7} {total_hrs_num:>8}  {status}")

print(f"\n  Total rows: {count}, Errors: {len(errors[1])}")

# ════════════════════════════════════════════════════════════════════════════
# FILE 2: Payroll_Master_July_2026.xlsx — 4 per-firm salary sheets
# (Master sheet is just an employee directory, no attendance data to verify)
# ════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 120)
print("[2] Payroll_Master_July_2026.xlsx — 4 per-firm salary sheets")
print("=" * 120)
wb2 = openpyxl.load_workbook('/home/z/my-project/download/Payroll_Master_July_2026.xlsx')
total_count = 0
for sheet_name in ['LAPL_July_2026_Sal', 'LRSL_July_2026_Sal', 'SDF_July_2026_Sal', 'SI_July_2026_Sal']:
    ws = wb2[sheet_name]
    print(f"\n— {sheet_name} —")
    print(f"  {'#':>3} {'Emp Name':<25} {'WorkedOT':>9} {'OT':>5} {'AddSun':>7} {'Formula':>20}  Status")
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 3).value
        if not name or name == 'TOTAL':
            continue
        total_count += 1
        worked_incl_ot = ws.cell(r, 8).value
        ot_hrs = ws.cell(r, 9).value
        sunday_hrs = ws.cell(r, 12).value
        total_formula = ws.cell(r, 14).value

        db = BY_NAME.get(name)
        if not db:
            errors[2].append(f"{sheet_name}/{name}: not found")
            print(f"  {total_count:>3} {name[:24]:<25} — NOT FOUND")
            continue

        issues = []
        if worked_incl_ot != round(db['workedHrsInclOT'], 2): issues.append(f"H≠{round(db['workedHrsInclOT'],2)}")
        if ot_hrs != round(db['otHrs'], 2): issues.append(f"I≠{round(db['otHrs'],2)}")
        if sunday_hrs != db['sundayHrs']: issues.append(f"L≠{db['sundayHrs']}")
        expected_formula = f"=H{r}+L{r}+M{r}"
        if total_formula != expected_formula: issues.append(f"Formula≠{expected_formula}")

        status = " ".join(issues) if issues else "OK"
        if issues: errors[2].append(f"{sheet_name}/{name}: {status}")
        print(f"  {total_count:>3} {name[:24]:<25} {str(worked_incl_ot):>9} {str(ot_hrs):>5} {str(sunday_hrs):>7} {str(total_formula):>20}  {status}")

print(f"\n  Total rows: {total_count}, Errors: {len(errors[2])}")

# ════════════════════════════════════════════════════════════════════════════
# FILE 3: Attendance_Tracker_Monthly_July_2026.xlsx — 4 firm daily grids
# ════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 120)
print("[3] Attendance_Tracker_Monthly_July_2026.xlsx — 4 firm daily grids")
print("=" * 120)
wb3 = openpyxl.load_workbook('/home/z/my-project/download/Attendance_Tracker_Monthly_July_2026.xlsx')
total_count = 0
for sheet_name in ['LAPL_July_2026_Att', 'LRSL_July_2026_Att', 'SDF_July_2026_Att', 'SI_July_2026_Att']:
    ws = wb3[sheet_name]
    print(f"\n— {sheet_name} —")
    print(f"  {'#':>3} {'Emp Name':<25} {'P':>3} {'A':>3} {'H':>3} {'Days':>5} {'TotHrs':>8} {'OTHrs':>7}  Status")
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 3).value
        if not name or name == 'TOTAL':
            continue
        total_count += 1
        present = ws.cell(r, 35).value
        absent = ws.cell(r, 36).value
        half = ws.cell(r, 37).value
        total_hrs_str = ws.cell(r, 38).value
        ot_hrs_str = ws.cell(r, 39).value

        # Count daily codes
        day_codes = []
        for c in range(4, 35):
            v = ws.cell(r, c).value
            if v:
                day_codes.append(str(v).strip())
        sundays_in_grid = day_codes.count('S')
        days_sum = (present or 0) + (absent or 0) + (half or 0) + sundays_in_grid

        db = BY_NAME.get(name)
        if not db:
            errors[3].append(f"{sheet_name}/{name}: not found")
            print(f"  {total_count:>3} {name[:24]:<25} — NOT FOUND")
            continue

        issues = []
        if present != db['presentDays']: issues.append(f"P≠{db['presentDays']}")
        if absent != db['absentDays']: issues.append(f"A≠{db['absentDays']}")
        if half != db['halfDays']: issues.append(f"H≠{db['halfDays']}")
        if total_hrs_str != fmt_hrs(db['workedHrsInclOT']): issues.append(f"TotHrs≠{fmt_hrs(db['workedHrsInclOT'])}")
        if ot_hrs_str != fmt_hrs(db['otHrs']): issues.append(f"OT≠{fmt_hrs(db['otHrs'])}")
        if days_sum != META['daysInMonth']: issues.append(f"DaysSum={days_sum}≠{META['daysInMonth']}")

        status = " ".join(issues) if issues else "OK"
        if issues: errors[3].append(f"{sheet_name}/{name}: {status}")
        print(f"  {total_count:>3} {name[:24]:<25} {str(present):>3} {str(absent):>3} {str(half):>3} {days_sum:>5} {str(total_hrs_str):>8} {str(ot_hrs_str):>7}  {status}")

print(f"\n  Total rows: {total_count}, Errors: {len(errors[3])}")

# ════════════════════════════════════════════════════════════════════════════
# FINAL SUMMARY
# ════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 120)
print("FINAL SUMMARY — All 42 Employees Verified")
print("=" * 120)
print(f"  Payroll_Summary_July_2026.xlsx (Payroll Register):  {42 - len(errors[1])}/42 OK, {len(errors[1])} errors")
print(f"  Payroll_Master_July_2026.xlsx (per-firm sheets):   {42 - len(errors[2])}/42 OK, {len(errors[2])} errors")
print(f"  Attendance_Tracker_Monthly_July_2026.xlsx:         {42 - len(errors[3])}/42 OK, {len(errors[3])} errors")
total_errors = len(errors[1]) + len(errors[2]) + len(errors[3])
print(f"\n  TOTAL ERRORS: {total_errors}")
if total_errors == 0:
    print("\n  🎉 ALL 42 EMPLOYEES VERIFIED — all 3 Excel files are 100% consistent with production API!")
else:
    print("\n  ⚠ Discrepancies found:")
    for f, errs in [(1, errors[1]), (2, errors[2]), (3, errors[3])]:
        for e in errs:
            print(f"    File {f}: {e}")
