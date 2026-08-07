"""Verify ALL 42 employees across:
1. Production DB (ground truth via july_data.json)
2. Production API /api/attendance/monthly-summary
3. Payroll_Summary_July_2026.xlsx — Payroll Register sheet
4. Payroll_Master_July_2026.xlsx — Master sheet + 4 firm salary sheets
5. Attendance_Tracker_Monthly_July_2026.xlsx — 4 firm attendance grids

Reports any inconsistencies.
"""
import json
import openpyxl
import urllib.request
import ssl

# Load ground truth from DB dump
DATA = json.load(open('/home/z/my-project/scripts/july_data.json'))
META = DATA['meta']
EMPLOYEES = {e['employeeId']: e for e in DATA['employees']}
print(f"Ground truth: {len(EMPLOYEES)} employees from DB")
print(f"  Days in month: {META['daysInMonth']}, Sundays: {META['numSundays']}, Holidays: {META['numHolidays']}, Working days: {META['workingDays']}")

# ── Pull production API data for every employee ──
print(f"\nFetching production API for {len(EMPLOYEES)} employees...")
api_data = {}
ctx = ssl.create_default_context()
for emp_id in EMPLOYEES:
    url = f"https://hrms.laxree.com/api/attendance/monthly-summary?employeeId={emp_id}&month=7&year=2026"
    try:
        with urllib.request.urlopen(url, context=ctx, timeout=30) as r:
            d = json.loads(r.read())
            api_data[emp_id] = d
    except Exception as e:
        print(f"  ⚠ {emp_id}: {e}")
print(f"  Got API data for {len(api_data)} employees")

# ── Helper: format hours ──
def fmt_hrs(v):
    if v is None or v == 0: return "0:00"
    h = int(v)
    m = int(round((v - h) * 60))
    if m == 60: h += 1; m = 0
    return f"{h}:{m:02d}"

# ── Verify File 1: Payroll_Summary_July_2026.xlsx ──
print("\n" + "=" * 110)
print("[1] Payroll_Summary_July_2026.xlsx — Payroll Register sheet")
print("=" * 110)
wb1 = openpyxl.load_workbook('/home/z/my-project/download/Payroll_Summary_July_2026.xlsx', data_only=False)
ws1 = wb1['Payroll Register']
print(f"{'Emp Name':<25} {'Present':>7} {'Absent':>7} {'WorkedOT':>9} {'AddSun':>7} {'Total':>8} {'Gross':>10}  Status")
print("-" * 110)
errors_1 = []
for r in range(5, ws1.max_row + 1):
    name = ws1.cell(r, 2).value
    if not name or name == 'TOTAL':
        continue
    present = ws1.cell(r, 6).value
    absent = ws1.cell(r, 7).value
    worked_hrs_str = ws1.cell(r, 8).value  # HH:MM string
    add_hrs_str = ws1.cell(r, 9).value     # HH:MM string
    total_hrs_num = ws1.cell(r, 10).value  # numeric
    gross_formula = ws1.cell(r, 11).value

    # Find this employee in DB by name
    db_emp = None
    for e in EMPLOYEES.values():
        if e['fullName'] == name:
            db_emp = e
            break
    if not db_emp:
        print(f"  ⚠ {name}: NOT FOUND in DB!")
        errors_1.append(f"{name}: not in DB")
        continue

    # Verify each value
    status_parts = []
    if present != db_emp['presentDays']:
        status_parts.append(f"Present≠{db_emp['presentDays']}")
    if absent != db_emp['absentDays']:
        status_parts.append(f"Absent≠{db_emp['absentDays']}")
    if worked_hrs_str != fmt_hrs(db_emp['workedHrsInclOT']):
        status_parts.append(f"WrkHrs≠{fmt_hrs(db_emp['workedHrsInclOT'])}")
    if add_hrs_str != fmt_hrs(db_emp['sundayHrs']):
        status_parts.append(f"Sun≠{fmt_hrs(db_emp['sundayHrs'])}")
    if total_hrs_num != round(db_emp['totalHrs'], 2):
        status_parts.append(f"Total≠{db_emp['totalHrs']}")
    # Days sum check: present + absent + sundays should = 31 (or 30 if half-day counted as 0.5)
    expected_total_days = present + absent + META['numSundays']
    if expected_total_days != META['daysInMonth']:
        status_parts.append(f"DaysSum={expected_total_days}≠{META['daysInMonth']}")

    status = " ".join(status_parts) if status_parts else "OK"
    if status_parts:
        errors_1.append(f"{name}: {status}")

    print(f"{name[:24]:<25} {str(present):>7} {str(absent):>7} {str(worked_hrs_str):>9} {str(add_hrs_str):>7} {str(total_hrs_num):>8} {'formula':>10}  {status}")

print(f"\nErrors in File 1: {len(errors_1)}")
for e in errors_1:
    print(f"  ✗ {e}")

# ── Verify File 2: Payroll_Master_July_2026.xlsx ──
print("\n" + "=" * 110)
print("[2] Payroll_Master_July_2026.xlsx — Master + 4 firm salary sheets")
print("=" * 110)
wb2 = openpyxl.load_workbook('/home/z/my-project/download/Payroll_Master_July_2026.xlsx', data_only=False)

# Master sheet check
print("\n— Master sheet —")
ws2m = wb2['Master']
print(f"{'Emp Name':<25} {'Present':>7} {'Absent':>7} {'WorkedOT':>9} {'Total':>8}  Status")
print("-" * 80)
errors_2_master = []
for r in range(5, ws2m.max_row + 1):
    name = ws2m.cell(r, 3).value
    if not name or name == 'TOTAL':
        continue
    present = ws2m.cell(r, 9).value
    absent = ws2m.cell(r, 10).value
    worked_hrs_str = ws2m.cell(r, 11).value
    total_hrs_str = ws2m.cell(r, 12).value

    db_emp = None
    for e in EMPLOYEES.values():
        if e['fullName'] == name:
            db_emp = e
            break
    if not db_emp:
        print(f"  ⚠ {name}: NOT FOUND in DB!")
        errors_2_master.append(f"{name}: not in DB")
        continue

    status_parts = []
    if present != db_emp['presentDays']:
        status_parts.append(f"Present≠{db_emp['presentDays']}")
    if absent != db_emp['absentDays']:
        status_parts.append(f"Absent≠{db_emp['absentDays']}")
    if worked_hrs_str != fmt_hrs(db_emp['workedHrsInclOT']):
        status_parts.append(f"WrkHrs≠{fmt_hrs(db_emp['workedHrsInclOT'])}")
    if total_hrs_str != fmt_hrs(db_emp['totalHrs']):
        status_parts.append(f"Total≠{fmt_hrs(db_emp['totalHrs'])}")

    status = " ".join(status_parts) if status_parts else "OK"
    if status_parts:
        errors_2_master.append(f"{name}: {status}")
    print(f"{name[:24]:<25} {str(present):>7} {str(absent):>7} {str(worked_hrs_str):>9} {str(total_hrs_str):>8}  {status}")

# Per-firm salary sheets check
print(f"\n— Per-firm salary sheets —")
errors_2_firm = []
for sheet_name in ['LAPL_July_2026_Sal', 'LRSL_July_2026_Sal', 'SDF_July_2026_Sal', 'SI_July_2026_Sal']:
    ws = wb2[sheet_name]
    print(f"\n  Sheet: {sheet_name}")
    print(f"  {'Emp Name':<25} {'WorkedOT':>9} {'OT':>5} {'AddSun':>7} {'TotalFormula':>20}  Status")
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 3).value
        if not name or name == 'TOTAL':
            continue
        worked_incl_ot = ws.cell(r, 8).value  # H — numeric
        ot_hrs = ws.cell(r, 9).value          # I — numeric
        sunday_hrs = ws.cell(r, 12).value     # L — numeric
        total_formula = ws.cell(r, 14).value  # N — formula string

        db_emp = None
        for e in EMPLOYEES.values():
            if e['fullName'] == name:
                db_emp = e
                break
        if not db_emp:
            print(f"    ⚠ {name}: NOT FOUND in DB!")
            errors_2_firm.append(f"{sheet_name}/{name}: not in DB")
            continue

        status_parts = []
        if worked_incl_ot != round(db_emp['workedHrsInclOT'], 2):
            status_parts.append(f"H≠{round(db_emp['workedHrsInclOT'],2)}")
        if ot_hrs != round(db_emp['otHrs'], 2):
            status_parts.append(f"I≠{round(db_emp['otHrs'],2)}")
        if sunday_hrs != db_emp['sundayHrs']:
            status_parts.append(f"L≠{db_emp['sundayHrs']}")
        # Check Total Hrs formula = H+L+M (not H+I+L+M)
        expected_formula = f"=H{r}+L{r}+M{r}"
        if total_formula != expected_formula:
            status_parts.append(f"Formula≠{expected_formula}")

        status = " ".join(status_parts) if status_parts else "OK"
        if status_parts:
            errors_2_firm.append(f"{sheet_name}/{name}: {status}")
        print(f"  {name[:24]:<25} {str(worked_incl_ot):>9} {str(ot_hrs):>5} {str(sunday_hrs):>7} {str(total_formula):>20}  {status}")

print(f"\nErrors in File 2 (Master): {len(errors_2_master)}")
for e in errors_2_master:
    print(f"  ✗ {e}")
print(f"Errors in File 2 (Per-firm sheets): {len(errors_2_firm)}")
for e in errors_2_firm:
    print(f"  ✗ {e}")

# ── Verify File 3: Attendance_Tracker_Monthly_July_2026.xlsx ──
print("\n" + "=" * 110)
print("[3] Attendance_Tracker_Monthly_July_2026.xlsx — 4 firm attendance grids")
print("=" * 110)
wb3 = openpyxl.load_workbook('/home/z/my-project/download/Attendance_Tracker_Monthly_July_2026.xlsx', data_only=False)
errors_3 = []
for sheet_name in ['LAPL_July_2026_Att', 'LRSL_July_2026_Att', 'SDF_July_2026_Att', 'SI_July_2026_Att']:
    ws = wb3[sheet_name]
    print(f"\n— Sheet: {sheet_name} ({ws.max_row - 4} employee rows) —")
    print(f"  {'Emp Name':<25} {'Present':>7} {'Absent':>7} {'Half':>5} {'TotalHrs':>9} {'OTHrs':>7}  Status")
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 3).value
        if not name or name == 'TOTAL':
            continue
        # Summary cols are at the end: 35=Present, 36=Absent, 37=Half, 38=Total Hrs, 39=OT Hrs
        present = ws.cell(r, 35).value
        absent = ws.cell(r, 36).value
        half = ws.cell(r, 37).value
        total_hrs_str = ws.cell(r, 38).value
        ot_hrs_str = ws.cell(r, 39).value

        # Count P/A/H/S in daily grid (cols 4 to 34, days 1-31)
        day_codes = []
        for c in range(4, 35):
            v = ws.cell(r, c).value
            if v:
                day_codes.append(str(v).strip())

        db_emp = None
        for e in EMPLOYEES.values():
            if e['fullName'] == name:
                db_emp = e
                break
        if not db_emp:
            print(f"    ⚠ {name}: NOT FOUND in DB!")
            errors_3.append(f"{sheet_name}/{name}: not in DB")
            continue

        status_parts = []
        if present != db_emp['presentDays']:
            status_parts.append(f"P≠{db_emp['presentDays']}")
        if absent != db_emp['absentDays']:
            status_parts.append(f"A≠{db_emp['absentDays']}")
        if half != db_emp['halfDays']:
            status_parts.append(f"H≠{db_emp['halfDays']}")
        # Total hrs in attendance tracker = worked hrs incl OT (not including Sunday)
        if total_hrs_str != fmt_hrs(db_emp['workedHrsInclOT']):
            status_parts.append(f"TotHrs≠{fmt_hrs(db_emp['workedHrsInclOT'])}")
        # OT hrs
        if ot_hrs_str != fmt_hrs(db_emp['otHrs']):
            status_parts.append(f"OT≠{fmt_hrs(db_emp['otHrs'])}")
        # Days sum: P + A + H + Sundays(4) should = 31
        sundays_in_grid = day_codes.count('S')
        days_sum = present + absent + half + sundays_in_grid
        if days_sum != META['daysInMonth']:
            status_parts.append(f"DaysSum={days_sum}≠{META['daysInMonth']}")

        status = " ".join(status_parts) if status_parts else "OK"
        if status_parts:
            errors_3.append(f"{sheet_name}/{name}: {status}")
        print(f"  {name[:24]:<25} {str(present):>7} {str(absent):>7} {str(half):>5} {str(total_hrs_str):>9} {str(ot_hrs_str):>7}  {status}")

print(f"\nErrors in File 3: {len(errors_3)}")
for e in errors_3:
    print(f"  ✗ {e}")

# ── Verify File 4: Production API ──
print("\n" + "=" * 110)
print("[4] Production API /api/attendance/monthly-summary — all 42 employees")
print("=" * 110)
print(f"{'EmpID':<10} {'Name':<25} {'P':>3} {'A':>3} {'Sun':>4} {'SunHrs':>7} {'TotHrs':>8}  Status")
print("-" * 80)
errors_4 = []
for emp_id, emp in EMPLOYEES.items():
    api = api_data.get(emp_id)
    if not api:
        print(f"  ⚠ {emp_id} ({emp['fullName']}): NO API DATA")
        errors_4.append(f"{emp_id}: no API data")
        continue
    p = api.get('presentDays')
    a = api.get('absentDays')
    sun = api.get('sundays')
    sun_hrs = api.get('totalSundayHours')
    tot_hrs = api.get('totalHrs')

    status_parts = []
    if p != emp['presentDays']:
        status_parts.append(f"P≠{emp['presentDays']}")
    if a != emp['absentDays']:
        status_parts.append(f"A≠{emp['absentDays']}")
    if sun != emp['numSundays']:
        status_parts.append(f"Sun≠{emp['numSundays']}")
    if sun_hrs != emp['sundayHrs']:
        status_parts.append(f"SunHrs≠{emp['sundayHrs']}")
    if round(tot_hrs or 0, 2) != round(emp['totalHrs'], 2):
        status_parts.append(f"TotHrs≠{emp['totalHrs']}")

    status = " ".join(status_parts) if status_parts else "OK"
    if status_parts:
        errors_4.append(f"{emp_id} ({emp['fullName']}): {status}")
    print(f"{emp_id:<10} {emp['fullName'][:24]:<25} {str(p):>3} {str(a):>3} {str(sun):>4} {str(sun_hrs):>7} {str(tot_hrs):>8}  {status}")

# ── Summary ──
print("\n" + "=" * 110)
print("FINAL SUMMARY")
print("=" * 110)
total_employees = len(EMPLOYEES)
print(f"Total employees verified: {total_employees}")
print(f"Errors in Payroll_Summary_July_2026.xlsx: {len(errors_1)}")
print(f"Errors in Payroll_Master_July_2026.xlsx (Master sheet): {len(errors_2_master)}")
print(f"Errors in Payroll_Master_July_2026.xlsx (per-firm sheets): {len(errors_2_firm)}")
print(f"Errors in Attendance_Tracker_Monthly_July_2026.xlsx: {len(errors_3)}")
print(f"Errors in Production API: {len(errors_4)}")
total_errors = len(errors_1) + len(errors_2_master) + len(errors_2_firm) + len(errors_3) + len(errors_4)
print(f"\nTOTAL ERRORS: {total_errors}")
if total_errors == 0:
    print("\n🎉 ALL 42 EMPLOYEES VERIFIED — all 3 Excel files + production API are consistent!")
else:
    print(f"\n⚠ {total_errors} discrepancies found — see details above.")
