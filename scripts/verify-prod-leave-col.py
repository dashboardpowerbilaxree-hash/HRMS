#!/usr/bin/env python3
"""
Verify all 42 employees' Leave column in PRODUCTION export-master Excel.
Cross-check against production DB ground truth.
"""
import psycopg2
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import load_workbook

DB_URL = open("/tmp/db_url.txt").read().strip()
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

# Get all employees with shift hours
cur.execute("""
    SELECT "employeeId", "fullName", firm, "shiftHours", "shiftStart", "shiftEnd"
    FROM "Employee" WHERE status = 'Yes' ORDER BY "fullName"
""")
emps = {r[0]: {'code': r[0], 'name': r[1], 'firm': r[2], 'shiftHours': r[3] or 9,
               'shiftStart': r[4], 'shiftEnd': r[5]} for r in cur.fetchall()}

# Get all July 2026 attendance
cur.execute("""
    SELECT "employeeId", date, status, "checkIn", "checkOut", "totalHours",
           "isSunday", "isHoliday", "halfDay", "lateEntry", "earlyOut", "overtimeHours"
    FROM "Attendance" WHERE date >= '2026-07-01' AND date <= '2026-07-31'
""")
att = defaultdict(list)
for r in cur.fetchall():
    att[r[0]].append({'day': r[1].day, 'status': r[2], 'checkIn': r[3], 'checkOut': r[4],
                      'totalHours': r[5] or 0, 'isSunday': r[6], 'isHoliday': r[7],
                      'halfDay': r[8], 'lateEntry': r[9], 'earlyOut': r[10], 'overtimeHours': r[11] or 0})

# Get all July 2026 leaves
cur.execute("""
    SELECT "employeeId", "startDate", "endDate" FROM "Leave"
    WHERE "startDate" <= '2026-07-31' AND "endDate" >= '2026-07-01' AND status = 'approved'
""")
lv_dates = defaultdict(set)
for ec, sd, ed in cur.fetchall():
    c = sd.date() if hasattr(sd,'date') else sd
    e = ed.date() if hasattr(ed,'date') else ed
    while c <= e:
        if date(2026,7,1) <= c <= date(2026,7,31):
            lv_dates[ec].add(c.isoformat())
        c += timedelta(days=1)

cur.close(); conn.close()

# Production recomputeStatus from src/lib/payroll-calc.ts
def time_str_to_minutes(t):
    if not t: return None
    parts = str(t).split(':')
    if len(parts) != 2: return None
    try: return int(parts[0]) * 60 + int(parts[1])
    except: return None

def get_actual_shift_hours(stored, shift_start, shift_end):
    """Compute actual shift hours from shift start/end, falling back to stored."""
    if shift_start and shift_end:
        s = time_str_to_minutes(shift_start)
        e = time_str_to_minutes(shift_end)
        if s is not None and e is not None:
            diff = e - s
            if diff < 0: diff += 24 * 60  # overnight
            return diff / 60.0
    return stored or 9

def is_actually_half_day(total_hours, actual_shift_hours):
    """Production: half-day if totalHours < 75% of shift (was 50%, now 75%)."""
    if not total_hours or total_hours <= 0: return False
    threshold = actual_shift_hours * 0.75
    return total_hours < threshold

def is_actually_early_out(check_out, shift_start, shift_end):
    """Production: early-out if checkout < shift end - 5 min grace."""
    if not check_out or not shift_start or not shift_end: return False
    co = time_str_to_minutes(check_out)
    ss = time_str_to_minutes(shift_start)
    se = time_str_to_minutes(shift_end)
    if co is None or ss is None or se is None: return False
    # Handle overnight shifts
    if se < ss: se += 24 * 60
    if co < ss: co += 24 * 60
    return co < se - 5

def recompute_status(rec, actual_shift_hours, shift_start, shift_end):
    """Production recomputeStatus from payroll-calc.ts."""
    is_stored_half_day = rec['status'] in ('half-day','half_day') or rec.get('halfDay')
    if is_stored_half_day:
        if is_actually_half_day(rec['totalHours'], actual_shift_hours):
            return 'half-day'
        actually_early_out = is_actually_early_out(rec['checkOut'], shift_start, shift_end) if (shift_start and shift_end) else rec.get('earlyOut')
        if rec.get('lateEntry'): return 'late'
        if actually_early_out: return 'early-out'
        return 'present'
    is_stored_early_out = rec['status'] == 'early-out' or rec.get('earlyOut')
    if is_stored_early_out and shift_start and shift_end:
        actually_early_out = is_actually_early_out(rec['checkOut'], shift_start, shift_end)
        if not actually_early_out:
            if rec.get('lateEntry'): return 'late'
            return 'present'
        return 'early-out'
    return rec['status']

# Compute ground truth for each employee (matches export-master logic)
YEAR, MONTH, DAYS = 2026, 7, 31
sundays = [d for d in range(1, DAYS+1) if date(YEAR, MONTH, d).weekday() == 6]
total_working_days = DAYS - len(sundays)

truth = {}
for ec, emp in emps.items():
    shift_hours = emp['shiftHours']
    actual_shift = get_actual_shift_hours(shift_hours, emp['shiftStart'], emp['shiftEnd'])
    attendance = att.get(ec, [])
    leave_days_set = lv_dates.get(ec, set())

    present_full = 0
    half = 0
    absent_days = 0
    leave_days = 0
    present_date_strs = set()

    # First pass: build present date strings
    for a in attendance:
        st = recompute_status(a, actual_shift, emp['shiftStart'], emp['shiftEnd'])
        if st in ('present','late','early-out','half-day','half_day'):
            ds = f"2026-07-{a['day']:02d}"
            present_date_strs.add(ds)

    # Second pass: count per export-master logic
    for d in range(1, DAYS+1):
        ds = f"2026-07-{d:02d}"
        date_obj = date(YEAR, MONTH, d)
        is_sunday = date_obj.weekday() == 6
        # find attendance record for this day
        rec = next((a for a in attendance if a['day'] == d), None)

        if rec:
            st = recompute_status(rec, actual_shift, emp['shiftStart'], emp['shiftEnd'])
            if st == 'absent':
                is_leave_day = ds in leave_days_set and ds not in present_date_strs
                if is_leave_day and not is_sunday:
                    leave_days += 1
                else:
                    absent_days += 1
            elif st == 'weekly-off':
                if rec['checkIn'] and rec['totalHours'] > 0:
                    present_full += 1
            elif st == 'holiday':
                if rec['checkIn'] and rec['totalHours'] > 0:
                    present_full += 1
            elif st == 'half-day' or st == 'half_day':
                present_full += 0.5
                absent_days += 0.5
                half += 1
            else:
                present_full += 1
        else:
            # No record
            is_leave_day = ds in leave_days_set and ds not in present_date_strs
            if is_sunday:
                pass
            elif is_leave_day:
                leave_days += 1
            else:
                absent_days += 1

    truth[ec] = {
        'name': emp['name'],
        'firm': emp['firm'],
        'present_full': present_full,
        'half': half,
        'absent_days': absent_days,
        'leave_days': leave_days,
        'total_leave_col': absent_days + leave_days,  # NEW: my fix
    }

# Load production Excel and extract Leave column for each employee
wb = load_workbook('/tmp/prod-master.xlsx', data_only=False)
firm_to_sheet = {'LAPL': 'LAPL', 'LRSL': 'LRSL', 'SI': 'SI', 'SDF': 'SDF'}

prod_leave = {}
for firm, sheet_name in firm_to_sheet.items():
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    # Section 3 starts after section 2. Find by detecting the pattern.
    # Each section has employee names in col 1. Section 3 ends with Total Working Hours + Leave in last 2 cols.
    # Section 3 rows: scan from bottom up to find rows with employee names
    for r in range(1, ws.max_row+1):
        name = ws.cell(r, 1).value
        if not name or name == 'EMP':
            continue
        # Check if this row has the Leave column (last col)
        leave_val = ws.cell(r, ws.max_column).value
        twh_val = ws.cell(r, ws.max_column - 1).value
        if leave_val is not None and twh_val is not None:
            # This is a section 3 row
            # Find this employee
            for ec, emp in emps.items():
                if emp['name'].lower() == str(name).lower():
                    prod_leave[ec] = leave_val
                    break

# Compare
print(f"{'#':3} {'EmpCode':10} {'Name':28} {'Firm':5} {'P':5} {'H':3} {'A_db':6} {'Lv_db':5} {'Sum_db':7} {'Prod_Leave':11} {'Match':6}")
print("-" * 110)
errors = []
for i, (ec, t) in enumerate(sorted(truth.items(), key=lambda x: x[1]['name']), 1):
    prod_lv = prod_leave.get(ec)
    expected = t['total_leave_col']
    match = "OK" if prod_lv == expected else "MISMATCH"
    if match == "MISMATCH":
        errors.append((ec, t['name'], expected, prod_lv))
    print(f"{i:3} {ec:10} {t['name'][:28]:28} {t['firm'] or '?':5} {t['present_full']:5.1f} {t['half']:3d} {t['absent_days']:6.2f} {t['leave_days']:5d} {t['present_full']+t['absent_days']+t['leave_days']+len(sundays):7.2f} {str(prod_lv):11} {match}")

print(f"\n{'='*110}")
print(f"Total: {len(truth)} employees, Mismatches: {len(errors)}")
if errors:
    print("\nMISMATCHES:")
    for ec, nm, exp, prod in errors:
        print(f"  {ec} - {nm}: expected={exp}, production={prod}")
