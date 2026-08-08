#!/usr/bin/env python3
"""
Verify all 42 employees' Leave column in PRODUCTION export-master Excel
against DB ground truth (using production recomputeStatus logic).
"""
import psycopg2
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import load_workbook

DB_URL = open("/tmp/db_url.txt").read().strip()
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

cur.execute("""
    SELECT "employeeId", "fullName", firm, "shiftHours", "shiftStart", "shiftEnd"
    FROM "Employee" WHERE status = 'Yes' ORDER BY "fullName"
""")
emps = {r[0]: {'code': r[0], 'name': r[1], 'firm': r[2], 'shiftHours': r[3] or 9,
               'shiftStart': r[4], 'shiftEnd': r[5]} for r in cur.fetchall()}

cur.execute("""
    SELECT "employeeId", date, status, "checkIn", "checkOut", "totalHours",
           "isSunday", "isHoliday", "halfDay", "lateEntry", "earlyOut", "overtimeHours"
    FROM "Attendance" WHERE date >= '2026-07-01' AND date <= '2026-07-31'
""")
att = defaultdict(list)
for r in cur.fetchall():
    att[r[0]].append({'day': r[1].day, 'status': r[2], 'checkIn': r[3], 'checkOut': r[4],
                      'totalHours': r[5] or 0, 'isSunday': r[6], 'isHoliday': r[7],
                      'halfDay': r[8], 'lateEntry': r[9], 'earlyOut': r[10], 'overtimeHours': r[11] or 0})

cur.execute("""
    SELECT "employeeId", "startDate", "endDate" FROM "Leave"
    WHERE "startDate" <= '2026-07-31' AND "endDate" >= '2026-07-01' AND status = 'approved'
""")
lv_dates = defaultdict(set)
for ec, sd, ed in cur.fetchall():
    c = sd.date() if hasattr(sd,'date') else sd
    e = ed.date() if hasattr(ed,'date') else ed
    while c <= e:
        if date(2026,7,1) <= c <= date(2026,7,31):
            lv_dates[ec].add(c.isoformat())
        c += timedelta(days=1)

cur.close(); conn.close()

# Production recomputeStatus from src/lib/payroll-calc.ts
def t2m(t):
    if not t: return None
    p = str(t).split(':')
    if len(p) != 2: return None
    try: return int(p[0]) * 60 + int(p[1])
    except: return None

def actual_shift(stored, ss, se):
    if ss and se:
        s, e = t2m(ss), t2m(se)
        if s is not None and e is not None:
            d = e - s
            if d < 0: d += 24*60
            return d / 60.0
    return stored or 9

def is_half(th, ash):
    # PRODUCTION: (totalHours || 0) < actualShiftHours / 2
    # Note: 0 hours IS considered a half-day (returns True for th=0)
    return (th or 0) < ash / 2

def is_early(co, ss, se):
    if not co or not ss or not se: return False
    co, ss, se = t2m(co), t2m(ss), t2m(se)
    if co is None or ss is None or se is None: return False
    if se < ss: se += 24*60
    if co < ss: co += 24*60
    return co < se - 5

def rstat(rec, ash, ss, se):
    is_sd_half = rec['status'] in ('half-day','half_day') or rec.get('halfDay')
    if is_sd_half:
        if is_half(rec['totalHours'], ash): return 'half-day'
        ae = is_early(rec['checkOut'], ss, se) if (ss and se) else rec.get('earlyOut')
        if rec.get('lateEntry'): return 'late'
        if ae: return 'early-out'
        return 'present'
    is_sd_eo = rec['status'] == 'early-out' or rec.get('earlyOut')
    if is_sd_eo and ss and se:
        ae = is_early(rec['checkOut'], ss, se)
        if not ae:
            if rec.get('lateEntry'): return 'late'
            return 'present'
        return 'early-out'
    return rec['status']

YEAR, MONTH, DAYS = 2026, 7, 31
sundays_count = 4

truth = {}
for ec, emp in emps.items():
    sh = emp['shiftHours']
    ash = actual_shift(sh, emp['shiftStart'], emp['shiftEnd'])
    attendance = att.get(ec, [])
    lds = lv_dates.get(ec, set())

    pf = 0; half = 0; absent_days = 0; leave_days = 0
    pds = set()
    for a in attendance:
        st = rstat(a, ash, emp['shiftStart'], emp['shiftEnd'])
        if st in ('present','late','early-out','half-day','half_day'):
            pds.add(f"2026-07-{a['day']:02d}")

    for d in range(1, DAYS+1):
        ds = f"2026-07-{d:02d}"
        is_sun = date(YEAR, MONTH, d).weekday() == 6
        rec = next((a for a in attendance if a['day'] == d), None)
        if rec:
            st = rstat(rec, ash, emp['shiftStart'], emp['shiftEnd'])
            if st == 'absent':
                ild = ds in lds and ds not in pds
                if ild and not is_sun: leave_days += 1
                else: absent_days += 1
            elif st == 'weekly-off':
                if rec['checkIn'] and rec['totalHours'] > 0: pf += 1
            elif st == 'holiday':
                if rec['checkIn'] and rec['totalHours'] > 0: pf += 1
            elif st in ('half-day','half_day'):
                pf += 0.5; absent_days += 0.5; half += 1
            else:
                pf += 1
        else:
            ild = ds in lds and ds not in pds
            if is_sun: pass
            elif ild: leave_days += 1
            else: absent_days += 1

    truth[ec] = {
        'name': emp['name'], 'firm': emp['firm'],
        'present': pf, 'half': half, 'absent': absent_days, 'leave': leave_days,
        'expected_leave_col': absent_days + leave_days,
    }

# Load production Excel
wb = load_workbook('/tmp/prod-master.xlsx', data_only=False)
prod = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    sections = []; cur_s = []
    for r in range(1, ws.max_row + 1):
        n = ws.cell(r, 1).value
        if n and n != 'EMP': cur_s.append(r)
        else:
            if cur_s: sections.append(cur_s); cur_s = []
    if cur_s: sections.append(cur_s)
    if len(sections) < 4: continue  # need at least 4 (title + 3 sections)
    sec3 = sections[-1]
    max_c = 0
    for r in sec3:
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value is not None:
                max_c = max(max_c, c)
    lv_col = max_c
    twh_col = max_c - 1
    for r in sec3:
        n = ws.cell(r, 1).value
        twh = ws.cell(r, twh_col).value
        lv = ws.cell(r, lv_col).value
        for ec, emp in emps.items():
            if emp['name'].lower() == str(n).lower():
                prod[ec] = {'twh': twh, 'leave': lv}
                break

# Compare
print(f"{'#':3} {'EmpCode':10} {'Name':26} {'Firm':5} {'P':5} {'H':3} {'A':6} {'Lv':3} {'Exp_Leave':10} {'Prod_Leave':11} {'Match':6}")
print("-" * 110)
errors = []
for i, (ec, t) in enumerate(sorted(truth.items(), key=lambda x: x[1]['name']), 1):
    pl = prod.get(ec, {}).get('leave')
    exp = t['expected_leave_col']
    # Normalize: production might store as int or float
    match = "OK"
    if pl is None:
        match = "NO_PROD"
        errors.append((ec, t['name'], exp, pl))
    elif abs(float(pl) - float(exp)) > 0.01:
        match = "MISMATCH"
        errors.append((ec, t['name'], exp, pl))
    print(f"{i:3} {ec:10} {t['name'][:26]:26} {t['firm'] or '?':5} {t['present']:5.1f} {t['half']:3d} {t['absent']:6.2f} {t['leave']:3d} {exp:10} {str(pl):11} {match}")

print(f"\n{'='*110}")
print(f"Total: {len(truth)} employees, Errors: {len(errors)}")
if errors:
    print("\nERRORS:")
    for ec, nm, exp, pl in errors:
        print(f"  {ec} - {nm}: expected={exp}, production={pl}")
else:
    print("\n✅ ALL 42 employees match between DB ground truth and production export-master Excel!")
    print(f"\nKamlesh verification: prod Leave = {prod.get('EMP-021', {}).get('leave')} (expected 3)")
