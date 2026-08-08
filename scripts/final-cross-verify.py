#!/usr/bin/env python3
"""FINAL verification: 42 employees × DB × 3 local Excel files × Production Excel"""
import psycopg2
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import load_workbook

DB_URL = open("/tmp/db_url.txt").read().strip()
conn = psycopg2.connect(DB_URL); cur = conn.cursor()

YEAR, MONTH, DAYS = 2026, 7, 31
SUNDAYS = [5, 12, 19, 26]

cur.execute("""SELECT "employeeId", "fullName", firm, "monthlySalary", "shiftHours", "shiftStart", "shiftEnd"
               FROM "Employee" WHERE status = 'Yes' ORDER BY "fullName" """)
emps = {r[0]: {'name': r[1], 'firm': r[2], 'salary': r[3] or 0, 'shift': r[4] or 9,
               'ss': r[5], 'se': r[6]} for r in cur.fetchall()}

cur.execute("""SELECT "employeeId", date, status, "checkIn", "checkOut", "totalHours",
               "isSunday", "isHoliday", "halfDay", "lateEntry", "earlyOut", "overtimeHours"
               FROM "Attendance" WHERE date >= '2026-07-01' AND date <= '2026-07-31'""")
att = defaultdict(list)
for r in cur.fetchall():
    att[r[0]].append({'day': r[1].day, 'status': r[2], 'ci': r[3], 'co': r[4],
                      'th': r[5] or 0, 'sun': r[6], 'hol': r[7], 'hd': r[8],
                      'le': r[9], 'eo': r[10], 'ot': r[11] or 0})

cur.execute("""SELECT "employeeId", "startDate", "endDate" FROM "Leave"
               WHERE "startDate" <= '2026-07-31' AND "endDate" >= '2026-07-01' AND status = 'approved'""")
lv = defaultdict(set)
for ec, sd, ed in cur.fetchall():
    c = sd.date() if hasattr(sd,'date') else sd
    e = ed.date() if hasattr(ed,'date') else ed
    while c <= e:
        if date(2026,7,1) <= c <= date(2026,7,31): lv[ec].add(c.isoformat())
        c += timedelta(days=1)
cur.close(); conn.close()

def t2m(t):
    if not t: return None
    p = str(t).split(':')
    if len(p) != 2: return None
    try: return int(p[0])*60 + int(p[1])
    except: return None

def ash(s, ss, se):
    if ss and se:
        a, b = t2m(ss), t2m(se)
        if a is not None and b is not None:
            d = b - a
            if d < 0: d += 24*60
            return d/60.0
    return s or 9

def is_h(th, a): return (th or 0) < a / 2

def is_e(co, ss, se):
    if not co or not ss or not se: return False
    co, ss, se = t2m(co), t2m(ss), t2m(se)
    if co is None or ss is None or se is None: return False
    if se < ss: se += 24*60
    if co < ss: co += 24*60
    return co < se - 5

def rs(r, a, ss, se):
    sh = r['status'] in ('half-day','half_day') or r.get('hd')
    if sh:
        if is_h(r['th'], a): return 'half-day'
        ae = is_e(r['co'], ss, se) if (ss and se) else r.get('eo')
        if r.get('le'): return 'late'
        if ae: return 'early-out'
        return 'present'
    se_eo = r['status'] == 'early-out' or r.get('eo')
    if se_eo and ss and se:
        ae = is_e(r['co'], ss, se)
        if not ae:
            if r.get('le'): return 'late'
            return 'present'
        return 'early-out'
    return r['status']

# DB ground truth
truth = {}
for ec, e in emps.items():
    a = ash(e['shift'], e['ss'], e['se'])
    at = att.get(ec, [])
    lds = lv.get(ec, set())
    pf=0; h=0; ad=0; ld=0; wh=0; pds=set()
    for r in at:
        s = rs(r, a, e['ss'], e['se'])
        if s in ('present','late','early-out','half-day','half_day'):
            pds.add(f"2026-07-{r['day']:02d}")
    for d in range(1, DAYS+1):
        ds = f"2026-07-{d:02d}"
        is_sun = date(YEAR, MONTH, d).weekday() == 6
        rec = next((x for x in at if x['day'] == d), None)
        if rec:
            s = rs(rec, a, e['ss'], e['se'])
            if s == 'absent':
                ild = ds in lds and ds not in pds
                if ild and not is_sun: ld += 1
                else: ad += 1
            elif s == 'weekly-off':
                if rec['ci'] and rec['th'] > 0: wh += rec['th']; pf += 1
            elif s == 'holiday':
                if rec['ci'] and rec['th'] > 0: wh += rec['th']; pf += 1
            elif s in ('half-day','half_day'):
                wh += rec['th']; pf += 0.5; ad += 0.5; h += 1
            else:
                wh += rec['th']; pf += 1
        else:
            ild = ds in lds and ds not in pds
            if is_sun: pass
            elif ild: ld += 1
            else: ad += 1
    sun_hrs = len(SUNDAYS) * e['shift']
    truth[ec] = {'name': e['name'], 'firm': e['firm'],
                 'P': pf, 'H': h, 'A': ad, 'Lv': ld, 'Leave': ad+ld,
                 'Worked': wh, 'Sun': sun_hrs, 'Total': wh+sun_hrs,
                 'salary': e['salary'], 'shift': e['shift'],
                 'slphr': e['salary']/(DAYS*e['shift']) if e['shift'] else 0,
                 'gross': (wh+sun_hrs) * (e['salary']/(DAYS*e['shift']) if e['shift'] else 0)}

# Load local Payroll_Summary (Payroll Register sheet)
wb1 = load_workbook('/home/z/my-project/download/Payroll_Summary_July_2026.xlsx')
ws1 = wb1['Payroll Register']
local_sum = {}
for r in range(5, ws1.max_row+1):
    nm = ws1.cell(r, 2).value
    if not nm: continue
    for ec, e in emps.items():
        if e['name'].lower() == str(nm).lower():
            local_sum[ec] = {
                'P': ws1.cell(r, 6).value, 'A': ws1.cell(r, 7).value,
                'Worked': ws1.cell(r, 8).value, 'Sun': ws1.cell(r, 9).value,
                'Total': ws1.cell(r, 10).value, 'gross': ws1.cell(r, 11).value,
            }

# Load local Payroll_Master (4 per-firm sheets) — extract Leave column from section 3
wb2 = load_workbook('/home/z/my-project/download/Payroll_Master_July_2026.xlsx')
local_master = {}
for sn in ['LAPL_July_2026_Sal','LRSL_July_2026_Sal','SDF_July_2026_Sal','SI_July_2026_Sal']:
    if sn not in wb2.sheetnames: continue
    ws = wb2[sn]
    sections = []; cs = []
    for r in range(1, ws.max_row+1):
        n = ws.cell(r, 1).value
        if n and n != 'EMP': cs.append(r)
        else:
            if cs: sections.append(cs); cs = []
    if cs: sections.append(cs)
    if len(sections) < 3: continue
    sec3 = sections[-1]
    max_c = 0
    for r in sec3:
        for c in range(1, ws.max_column+1):
            if ws.cell(r, c).value is not None:
                max_c = max(max_c, c)
    for r in sec3:
        n = ws.cell(r, 1).value
        for ec, e in emps.items():
            if e['name'].lower() == str(n).lower():
                local_master[ec] = {
                    'TWH': ws.cell(r, max_c-1).value,
                    'Leave': ws.cell(r, max_c).value,
                }

# Load production Excel (already in /tmp/prod-master.xlsx)
wb3 = load_workbook('/tmp/prod-master.xlsx')
prod = {}
for sn in wb3.sheetnames:
    ws = wb3[sn]
    sections = []; cs = []
    for r in range(1, ws.max_row+1):
        n = ws.cell(r, 1).value
        if n and n != 'EMP': cs.append(r)
        else:
            if cs: sections.append(cs); cs = []
    if cs: sections.append(cs)
    if len(sections) < 4: continue
    sec3 = sections[-1]
    max_c = 0
    for r in sec3:
        for c in range(1, ws.max_column+1):
            if ws.cell(r, c).value is not None:
                max_c = max(max_c, c)
    for r in sec3:
        n = ws.cell(r, 1).value
        for ec, e in emps.items():
            if e['name'].lower() == str(n).lower():
                prod[ec] = {'TWH': ws.cell(r, max_c-1).value, 'Leave': ws.cell(r, max_c).value}
                break

# Helper: parse "216:07" → 216.07 (decimal hours)
def parse_hrs(s):
    if s is None: return None
    if isinstance(s, (int, float)): return float(s)
    s = str(s)
    if ':' in s:
        parts = s.split(':')
        try: return int(parts[0]) + int(parts[1])/60
        except: return None
    try: return float(s)
    except: return None

# Compare
print(f"{'#':3} {'EmpCode':10} {'Name':24} | {'P_db':5} {'P_sum':5} | {'A_db':5} {'A_sum':5} | {'Lv_db':5} {'Lv_mas':6} {'Lv_prod':7} | {'T_db':7} {'T_sum':7} {'T_mas':7} {'T_prod':7} | Status")
print("-" * 160)
errors = 0
for i, (ec, t) in enumerate(sorted(truth.items(), key=lambda x: x[1]['name']), 1):
    ls = local_sum.get(ec, {})
    lm = local_master.get(ec, {})
    pr = prod.get(ec, {})

    issues = []
    # Present
    if ls.get('P') is not None and abs(float(ls['P']) - float(t['P'])) > 0.01:
        issues.append(f"P sum={ls['P']}≠{t['P']}")
    # Absent
    if ls.get('A') is not None and abs(float(ls['A']) - float(t['A'] + t['H']*0.5)) > 0.01:
        issues.append(f"A sum={ls['A']}≠{t['A']+t['H']*0.5}")
    # Leave (master + prod)
    if lm.get('Leave') is not None and abs(float(lm['Leave']) - float(t['Leave'])) > 0.01:
        issues.append(f"Leave mas={lm['Leave']}≠{t['Leave']}")
    if pr.get('Leave') is not None and abs(float(pr['Leave']) - float(t['Leave'])) > 0.01:
        issues.append(f"Leave prod={pr['Leave']}≠{t['Leave']}")
    # Total Hrs
    t_db = t['Total']
    t_sum = parse_hrs(ls.get('Total'))
    t_mas = parse_hrs(lm.get('TWH'))
    t_prod = parse_hrs(pr.get('TWH'))
    if t_sum is not None and abs(t_sum - t_db) > 0.05:
        issues.append(f"T sum={t_sum}≠{t_db}")
    if t_mas is not None and abs(t_mas - t_db) > 0.05:
        issues.append(f"T mas={t_mas}≠{t_db}")
    if t_prod is not None and abs(t_prod - t_db) > 0.05:
        issues.append(f"T prod={t_prod}≠{t_db}")

    status = "OK" if not issues else "MISMATCH: " + "; ".join(issues)
    if issues: errors += 1

    print(f"{i:3} {ec:10} {t['name'][:24]:24} | {t['P']:5.1f} {str(ls.get('P','')):5} | {t['A']:5.2f} {str(ls.get('A','')):5} | {t['Leave']:5} {str(lm.get('Leave','')):6} {str(pr.get('Leave','')):7} | {t_db:7.2f} {str(t_sum or ''):7} {str(t_mas or ''):7} {str(t_prod or ''):7} | {status}")

print(f"\n{'='*160}")
print(f"Total: {len(truth)} employees, Errors: {errors}")
if errors == 0:
    print("\n✅ ALL 42 EMPLOYEES 100% VERIFIED — DB = Local Summary = Local Master = Production Excel")
    print(f"\nKamlesh: P={truth['EMP-021']['P']}, A={truth['EMP-021']['A']}, Leave={truth['EMP-021']['Leave']}, Worked={int(truth['EMP-021']['Worked'])}:{round((truth['EMP-021']['Worked']-int(truth['EMP-021']['Worked']))*60):02d}, Total={int(truth['EMP-021']['Total'])}:{round((truth['EMP-021']['Total']-int(truth['EMP-021']['Total']))*60):02d}, Gross=₹{truth['EMP-021']['gross']:,.0f}")
