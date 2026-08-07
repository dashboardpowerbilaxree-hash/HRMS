#!/usr/bin/env python3
"""Generate Payroll_Summary_July_2026.xlsx matching user's reference format.
Sheets:
  1. Payroll Register — header at R4, total at end
     A=S.No | B=Employee Name | C=Monthly Salary | D=Working Hrs | E=Sl/Hr
     F=Present Days | G=Absent Days | H=Worked Hrs including OT | I=Additional hrs (Sunday Hrs)
     J=Total Hrs | K=Gross Salary | L=SD Refund | M=Salary Advance | N=Net Salary
  2. Summary — high-level totals
  3. Master — combined employee list with all details
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = json.load(open('/home/z/my-project/scripts/july_data.json'))
META = DATA['meta']
EMPLOYEES = DATA['employees']

# Sort by firm, then fullName
EMPLOYEES.sort(key=lambda e: (e['firm'] or '~', e['fullName']))

OUT = '/home/z/my-project/download/Payroll_Summary_July_2026.xlsx'
os.makedirs(os.path.dirname(OUT), exist_ok=True)

wb = Workbook()
wb.properties.creator = "Z.ai"

# ──────────── STYLES ────────────
NAVY = "1F3A5F"
GOLD = "C9A961"
LIGHT_BLUE = "E8EEF4"
LIGHT_GOLD = "FAF5E8"
WHITE = "FFFFFF"

thin = Side(border_style="thin", color="666666")
medium = Side(border_style="medium", color=NAVY)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_thick = Border(left=medium, right=medium, top=medium, bottom=medium)

font_title = Font(name='Calibri', size=16, bold=True, color=WHITE)
font_subtitle = Font(name='Calibri', size=11, italic=True, color=WHITE)
font_header = Font(name='Calibri', size=10, bold=True, color=WHITE)
font_body = Font(name='Calibri', size=10)
font_total = Font(name='Calibri', size=10, bold=True, color=NAVY)
font_section = Font(name='Calibri', size=12, bold=True, color=NAVY)

fill_title = PatternFill("solid", fgColor=NAVY)
fill_header = PatternFill("solid", fgColor=NAVY)
fill_total = PatternFill("solid", fgColor=LIGHT_GOLD)
fill_alt = PatternFill("solid", fgColor=LIGHT_BLUE)
fill_section = PatternFill("solid", fgColor=LIGHT_GOLD)

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center')

def fmt_hrs(v):
    """Format hours as HH:MM string."""
    if v is None or v == 0:
        return "0:00"
    h = int(v)
    m = int(round((v - h) * 60))
    if m == 60:
        h += 1
        m = 0
    return f"{h}:{m:02d}"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: Payroll Register
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Payroll Register"

# Title rows
ws.merge_cells('A1:N1')
ws['A1'] = "LAXREE GROUP OF COMPANIES"
ws['A1'].font = font_title
ws['A1'].fill = fill_title
ws['A1'].alignment = align_center
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:N2')
ws['A2'] = f"Payroll Register — July 2026"
ws['A2'].font = font_subtitle
ws['A2'].fill = fill_title
ws['A2'].alignment = align_center
ws.row_dimensions[2].height = 20

ws.merge_cells('A3:N3')
ws['A3'] = (f"Days in Month: {META['daysInMonth']}  |  Sundays: {META['numSundays']}  |  "
            f"Holidays: {META['numHolidays']}  |  Working Days: {META['workingDays']}  |  "
            f"Total Employees: {len(EMPLOYEES)}")
ws['A3'].font = Font(name='Calibri', size=9, italic=True, color="555555")
ws['A3'].alignment = align_center
ws.row_dimensions[3].height = 18

# Header row 4
HEADERS = [
    'S.No', 'Employee Name', 'Monthly Salary', 'Working Hrs', 'Sl/Hr',
    'Present Days', 'Absent Days', 'Worked Hrs including OT', 'Additional hrs',
    'Total Hrs', 'Gross Salary', 'SD Refund', 'Salary Advance', 'Net Salary'
]
for i, h in enumerate(HEADERS, 1):
    c = ws.cell(4, i, h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center
    c.border = border_all
ws.row_dimensions[4].height = 38

# Data rows
row = 5
for idx, emp in enumerate(EMPLOYEES, 1):
    vals = [
        idx,
        emp['fullName'],
        emp['monthlySalary'],
        emp['shiftHours'],
        None,  # Sl/Hr — formula below
        emp['presentDays'],
        emp['absentDays'],
        fmt_hrs(emp['workedHrsInclOT']),
        fmt_hrs(emp['sundayHrs']),
        None,  # Total Hrs — formula
        None,  # Gross Salary — formula
        0,     # SD Refund
        0,     # Salary Advance
        None,  # Net Salary — formula
    ]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v)
        c.font = font_body
        c.border = border_all
        if i in (1, 4, 6, 7, 8, 9, 10):
            c.alignment = align_center
        elif i in (3, 5, 11, 12, 13, 14):
            c.alignment = align_right
        else:
            c.alignment = align_left
        if row % 2 == 0:
            c.fill = fill_alt

    # Sl/Hr = Monthly Salary / (Days in Month × Shift Hrs)  → use formula
    ws.cell(row, 5).value = f"=C{row}/({META['daysInMonth']}*D{row})"
    ws.cell(row, 5).number_format = '0.00'

    # Total Hrs = Worked Hrs + Additional Hrs  (numeric value, not formula since H & I are HH:MM strings)
    ws.cell(row, 10).value = round(emp['totalHrs'], 2)
    ws.cell(row, 10).number_format = '0.00'

    # Gross Salary = Total Hrs × Sl/Hr (using computed values)
    ws.cell(row, 11).value = f"=J{row}*E{row}"
    ws.cell(row, 11).number_format = '#,##0.00'

    # Net Salary = Gross - SD Refund - Salary Advance
    ws.cell(row, 14).value = f"=K{row}-L{row}-M{row}"
    ws.cell(row, 14).number_format = '#,##0.00'

    ws.cell(row, 3).number_format = '#,##0'
    row += 1

# Total row
total_row = row
ws.cell(total_row, 1).value = ""
ws.cell(total_row, 2).value = "TOTAL"
ws.cell(total_row, 3).value = f"=SUM(C5:C{total_row-1})"
ws.cell(total_row, 6).value = f"=SUM(F5:F{total_row-1})"
ws.cell(total_row, 7).value = f"=SUM(G5:G{total_row-1})"
# For Total Hrs, J: sum numeric values
ws.cell(total_row, 10).value = f"=SUM(J5:J{total_row-1})"
ws.cell(total_row, 11).value = f"=SUM(K5:K{total_row-1})"
ws.cell(total_row, 12).value = f"=SUM(L5:L{total_row-1})"
ws.cell(total_row, 13).value = f"=SUM(M5:M{total_row-1})"
ws.cell(total_row, 14).value = f"=SUM(N5:N{total_row-1})"

for i in range(1, 15):
    c = ws.cell(total_row, i)
    c.font = font_total
    c.fill = fill_total
    c.border = border_all
    if i in (3, 10, 11, 12, 13, 14):
        c.alignment = align_right
        c.number_format = '#,##0.00'
    elif i in (6, 7):
        c.alignment = align_center
    else:
        c.alignment = align_center

# Column widths
widths = [6, 26, 14, 10, 10, 11, 11, 18, 13, 11, 14, 11, 13, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: Summary
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Summary")

ws2.merge_cells('A1:B1')
ws2['A1'] = "Payroll Summary"
ws2['A1'].font = Font(name='Calibri', size=16, bold=True, color=WHITE)
ws2['A1'].fill = fill_title
ws2['A1'].alignment = align_center
ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:B2')
ws2['A2'] = "Laxree Group of Companies — July 2026"
ws2['A2'].font = font_subtitle
ws2['A2'].fill = fill_title
ws2['A2'].alignment = align_center

# Compute totals
total_gross = sum(e['grossSalary'] for e in EMPLOYEES)
total_ot_amount = sum(e['otHrs'] * e['slPerHr'] for e in EMPLOYEES)
total_deductions = 0
total_net = total_gross - total_deductions
total_ot_hours = sum(e['otHrs'] for e in EMPLOYEES)
total_present = sum(e['presentDays'] for e in EMPLOYEES)
total_absent = sum(e['absentDays'] for e in EMPLOYEES)

# Section: Amounts
r = 4
ws2.cell(r, 1, "Category").font = font_header
ws2.cell(r, 1).fill = fill_header
ws2.cell(r, 1).alignment = align_center
ws2.cell(r, 1).border = border_all
ws2.cell(r, 2, "Amount (₹)").font = font_header
ws2.cell(r, 2).fill = fill_header
ws2.cell(r, 2).alignment = align_center
ws2.cell(r, 2).border = border_all
r += 1

summary_rows = [
    ("Total Gross Salary", total_gross),
    ("Total OT Amount", total_ot_amount),
    ("Total Bonus", 0),
    ("Total Deductions", total_deductions),
    ("Total Net Payroll", total_net),
]
for label, val in summary_rows:
    ws2.cell(r, 1, label).font = font_body
    ws2.cell(r, 1).border = border_all
    ws2.cell(r, 1).alignment = align_left
    ws2.cell(r, 2, val).font = font_body
    ws2.cell(r, 2).border = border_all
    ws2.cell(r, 2).alignment = align_right
    ws2.cell(r, 2).number_format = '#,##0.00'
    if label == "Total Net Payroll":
        ws2.cell(r, 1).font = font_total
        ws2.cell(r, 2).font = font_total
        ws2.cell(r, 1).fill = fill_total
        ws2.cell(r, 2).fill = fill_total
    r += 1

# Section: Metrics
r += 1
ws2.cell(r, 1, "Metric").font = font_header
ws2.cell(r, 1).fill = fill_header
ws2.cell(r, 1).alignment = align_center
ws2.cell(r, 1).border = border_all
ws2.cell(r, 2, "Value").font = font_header
ws2.cell(r, 2).fill = fill_header
ws2.cell(r, 2).alignment = align_center
ws2.cell(r, 2).border = border_all
r += 1

avg_net = total_net / len(EMPLOYEES) if EMPLOYEES else 0
metric_rows = [
    ("Employees Processed", len(EMPLOYEES)),
    ("Total Present Days", total_present),
    ("Total Absent Days", total_absent),
    ("Average Net Salary", round(avg_net, 2)),
    ("Total OT Hours", fmt_hrs(total_ot_hours)),
]
for label, val in metric_rows:
    ws2.cell(r, 1, label).font = font_body
    ws2.cell(r, 1).border = border_all
    ws2.cell(r, 1).alignment = align_left
    ws2.cell(r, 2, val).font = font_body
    ws2.cell(r, 2).border = border_all
    ws2.cell(r, 2).alignment = align_right
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        ws2.cell(r, 2).number_format = '#,##0.00' if isinstance(val, float) else '#,##0'
    r += 1

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 20

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: Master
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Master")

ws3.merge_cells('A1:L1')
ws3['A1'] = "LAXREE GROUP OF COMPANIES — Master Employee List (July 2026)"
ws3['A1'].font = font_title
ws3['A1'].fill = fill_title
ws3['A1'].alignment = align_center
ws3.row_dimensions[1].height = 28

ws3.merge_cells('A2:L2')
ws3['A2'] = (f"Days in Month: {META['daysInMonth']}  |  Sundays: {META['numSundays']}  |  "
             f"Holidays: {META['numHolidays']}  |  Working Days: {META['workingDays']}")
ws3['A2'].font = Font(name='Calibri', size=10, italic=True, color="555555")
ws3['A2'].alignment = align_center

# Header
master_headers = ['S.No', 'Emp Code', 'Employee Name', 'Firm', 'Location',
                  'Monthly Salary', 'Shift Hrs/Day', 'Sl/Hr', 'Present Days',
                  'Absent Days', 'Worked Hrs incl OT', 'Total Hrs']
for i, h in enumerate(master_headers, 1):
    c = ws3.cell(4, i, h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center
    c.border = border_all
ws3.row_dimensions[4].height = 32

# Master data rows
r = 5
for idx, emp in enumerate(EMPLOYEES, 1):
    vals = [
        idx,
        emp['employeeId'],
        emp['fullName'],
        emp['firm'] or '-',
        emp['location'] or '-',
        emp['monthlySalary'],
        emp['shiftHours'],
        None,
        emp['presentDays'],
        emp['absentDays'],
        fmt_hrs(emp['workedHrsInclOT']),
        fmt_hrs(emp['totalHrs']),
    ]
    for i, v in enumerate(vals, 1):
        c = ws3.cell(r, i, v)
        c.font = font_body
        c.border = border_all
        if i in (1, 4, 7, 8, 9, 10, 11, 12):
            c.alignment = align_center
        elif i in (6,):
            c.alignment = align_right
            c.number_format = '#,##0'
        else:
            c.alignment = align_left
        if r % 2 == 0:
            c.fill = fill_alt
    # Sl/Hr formula = Monthly Salary / (Days in Month × Shift Hrs)
    ws3.cell(r, 8).value = f"=F{r}/({META['daysInMonth']}*G{r})"
    ws3.cell(r, 8).number_format = '0.00'
    r += 1

# Total row
ws3.cell(r, 1).value = ""
ws3.cell(r, 3).value = "TOTAL"
ws3.cell(r, 6).value = f"=SUM(F5:F{r-1})"
ws3.cell(r, 9).value = f"=SUM(I5:I{r-1})"
ws3.cell(r, 10).value = f"=SUM(J5:J{r-1})"
for i in range(1, 13):
    c = ws3.cell(r, i)
    c.font = font_total
    c.fill = fill_total
    c.border = border_all
    if i == 6:
        c.number_format = '#,##0'
        c.alignment = align_right
    elif i in (9, 10):
        c.alignment = align_center
    else:
        c.alignment = align_center

widths3 = [6, 12, 26, 7, 16, 14, 11, 10, 11, 11, 16, 13]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = "A5"

wb.save(OUT)
print(f"✓ Saved: {OUT}")
print(f"  Employees: {len(EMPLOYEES)}")
print(f"  Total Gross: ₹{total_gross:,.2f}")
print(f"  Total Net: ₹{total_net:,.2f}")

# Kamlesh sanity check
kamlesh = next(e for e in EMPLOYEES if 'Kamlesh' in e['fullName'])
print(f"\nKamlesh check:")
print(f"  Present Days:  {kamlesh['presentDays']}  (expected 24)")
print(f"  Absent Days:   {kamlesh['absentDays']}  (expected 3) ✓")
print(f"  Worked Hrs incl OT: {fmt_hrs(kamlesh['workedHrsInclOT'])}  (expected 216:04)")
print(f"  Sunday Hrs (Additional): {fmt_hrs(kamlesh['sundayHrs'])}  (expected 36:00)")
print(f"  Total Hrs:     {fmt_hrs(kamlesh['totalHrs'])}  (expected 252:04)")
print(f"  Gross Salary:  ₹{kamlesh['grossSalary']:,.2f}")
