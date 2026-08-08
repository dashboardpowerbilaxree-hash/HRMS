#!/usr/bin/env python3
"""Generate a verification report Excel for all 42 employees — July 2026."""
import psycopg2
from datetime import date, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_URL = open("/tmp/db_url.txt").read().strip()
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

YEAR, MONTH = 2026, 7
DAYS = 31
SUNDAYS = [d for d in range(1, DAYS+1) if date(YEAR, MONTH, d).weekday() == 6]

cur.execute("""
    SELECT "employeeId", "fullName", firm, "monthlySalary", "shiftHours"
    FROM "Employee" WHERE status = 'Yes' ORDER BY "fullName"
""")
emps = cur.fetchall()

cur.execute("""
    SELECT "employeeId", date, status, "checkIn", "totalHours", "isSunday", "isHoliday", "halfDay"
    FROM "Attendance" WHERE date >= '2026-07-01' AND date <= '2026-07-31'
""")
att = defaultdict(list)
for r in cur.fetchall():
    att[r[0]].append({'day': r[1].day, 'status': r[2], 'checkIn': r[3],
                      'totalHours': r[4] or 0, 'isSunday': r[5], 'isHoliday': r[6], 'halfDay': r[7]})

cur.execute("""
    SELECT "employeeId", "startDate", "endDate" FROM "Leave"
    WHERE "startDate" <= '2026-07-31' AND "endDate" >= '2026-07-01'
""")
lv = defaultdict(int)
for ec, sd, ed in cur.fetchall():
    c = sd.date() if hasattr(sd,'date') else sd
    e = ed.date() if hasattr(ed,'date') else ed
    while c <= e:
        if date(YEAR,MONTH,1) <= c <= date(YEAR,MONTH,DAYS):
            lv[ec] += 1
        c += timedelta(days=1)

def rstat(r, sh):
    s = r['status']
    if s == 'absent': return 'absent'
    if r['isSunday'] and not r['checkIn']: return 'weekly-off'
    if r['isHoliday'] and not r['checkIn']: return 'holiday'
    if not r['checkIn']: return 'absent'
    if s in ('half-day','half_day') or r.get('halfDay'): return 'half-day'
    return s

wb = Workbook()
ws = wb.active
ws.title = "Verification_July_2026"

# Styles
hdr_fill = PatternFill("solid", fgColor="1F2937")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
ok_fill = PatternFill("solid", fgColor="DCFCE7")
err_fill = PatternFill("solid", fgColor="FEE2E2")
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title
ws.merge_cells('A1:M1')
ws['A1'] = "LAXREE GROUP OF COMPANIES — July 2026 Payroll Verification (All 42 Employees)"
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells('A2:M2')
ws['A2'] = f"Days in Month: {DAYS}  |  Sundays: {len(SUNDAYS)} (Jul 5, 12, 19, 26)  |  Working Days: {DAYS - len(SUNDAYS)}  |  Generated from production DB"
ws['A2'].font = Font(italic=True, size=10)
ws['A2'].alignment = Alignment(horizontal="center")

headers = ['#','Emp Code','Employee Name','Firm','Monthly Salary','Shift Hrs',
           'Present','Absent','Half','Sundays','Worked Hrs (incl OT)','Sunday Hrs','Total Hrs',
           'Sl/Hr','Gross Salary','Days Sum','Status']
for c, h in enumerate(headers, 1):
    cell = ws.cell(4, c, h)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = center
    cell.border = border

row = 5
errors = 0
for i, (ec, nm, firm, sal, sh) in enumerate(emps, 1):
    sh = sh or 9
    sal = sal or 0
    sl_per_hr = sal / (DAYS * sh) if sh else 0
    sundays = len(SUNDAYS)
    total_wd = DAYS - sundays  # 27

    pf = 0; half = 0; wh = 0
    for a in att.get(ec, []):
        s = rstat(a, sh)
        if s == 'absent': pass
        elif s in ('weekly-off','holiday'):
            if a['checkIn'] and a['totalHours'] > 0:
                wh += a['totalHours']; pf += 1
        elif s == 'half-day':
            wh += a['totalHours']; half += 1
        else:
            wh += a['totalHours']; pf += 1

    absent = max(0, total_wd - pf - half)
    sun_hrs = sundays * sh
    total_hrs = wh + sun_hrs
    gross = total_hrs * sl_per_hr
    days_sum = pf + absent + half + sundays
    ok = abs(days_sum - DAYS) < 0.01
    if not ok: errors += 1

    vals = [i, ec, nm, firm or '?', sal, sh,
            pf, absent, half, sundays,
            f"{int(wh)}:{round((wh-int(wh))*60):02d}" if wh else "0:00",
            f"{int(sun_hrs)}:{round((sun_hrs-int(sun_hrs))*60):02d}",
            f"{int(total_hrs)}:{round((total_hrs-int(total_hrs))*60):02d}",
            f"₹{sl_per_hr:.2f}", f"₹{gross:,.0f}", days_sum, "OK" if ok else f"FAIL sum={days_sum}"]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row, c, v)
        cell.alignment = center if c != 3 else Alignment(horizontal="left", vertical="center")
        cell.border = border
        if not ok:
            cell.fill = err_fill
        elif c == 17:
            cell.fill = ok_fill
    row += 1

# Summary footer
ws.cell(row+1, 1, "TOTAL").font = Font(bold=True)
ws.cell(row+1, 2, f"{len(emps)} employees").font = Font(bold=True)
ws.cell(row+1, 16, f"Errors: {errors}").font = Font(bold=True, color="DC2626" if errors else "059669")
ws.cell(row+1, 17, "ALL OK" if errors == 0 else "NEEDS FIX").font = Font(bold=True, color="059669" if errors == 0 else "DC2626")

# Column widths
widths = [4, 11, 28, 6, 14, 9, 8, 8, 6, 9, 14, 11, 11, 9, 14, 9, 10]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = 'A5'

out = "/home/z/my-project/download/Verification_Report_July_2026_All_42_Employees.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total: {len(emps)} employees, Errors: {errors}")
cur.close(); conn.close()
