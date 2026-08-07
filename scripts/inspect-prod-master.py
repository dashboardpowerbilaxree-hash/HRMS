"""Inspect export-master output."""
import openpyxl
wb = openpyxl.load_workbook('/tmp/test_master.xlsx', data_only=False)
print(f"Sheets: {wb.sheetnames}")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== Sheet: {sn} ({ws.max_row} rows x {ws.max_column} cols) ===")
    for r in range(1, min(6, ws.max_row + 1)):
        row = []
        for c in range(1, min(15, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            if v is None: v = ''
            row.append(str(v)[:16])
        print(f"  R{r}: {' | '.join(row)}")
    # find Kamlesh row
    for r in range(1, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value or ws.cell(row=r, column=3).value or ''
        if 'kamlesh' in str(name).lower():
            print(f"  KAMLESH @ row {r}:")
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                hdr = ws.cell(row=1, column=c).value or ws.cell(row=3, column=c).value or ws.cell(row=4, column=c).value or f"C{c}"
                print(f"    {hdr}: {v}")
            break
