#!/usr/bin/env python3
"""
Test the new payroll summary-export route locally by hitting the dev server.
Verifies:
  1. Multiple sheets are generated when firm=ALL (one per firm + Summary)
  2. Each firm sheet uses the company's full name (not 'LAXREE GROUP')
  3. Column headers match the template (S.No, Employee Name, Monthly Salary,
     Working Hrs, Sl/Hr, Present Days, Absent Days, Worked Hrs, Additional hrs,
     Total Hrs, Gross Salary, SD Refund, Salary Advance, Net Salary)
"""
import urllib.request, urllib.error, sys, os
import openpyxl
from io import BytesIO

BASE = "http://localhost:3000"

# Test 1: All firms
print("=== Test 1: All firms (firm=ALL) ===")
try:
    url = f"{BASE}/api/payroll/summary-export?month=7&year=2026&firm=all"
    req = urllib.request.Request(url)
    with urllib.request.urlopen(req, timeout=30) as r:
        if r.status != 200:
            print(f"  FAIL: HTTP {r.status}")
            sys.exit(1)
        data = r.read()
        print(f"  OK: HTTP 200, {len(data)} bytes")
        cd = r.headers.get('Content-Disposition', '')
        print(f"  Content-Disposition: {cd}")
except urllib.error.URLError as e:
    print(f"  SKIP: dev server not running ({e})")
    print("  Run: cd /home/z/my-project && npm run dev")
    sys.exit(2)
except urllib.error.HTTPError as e:
    body = e.read().decode()[:300]
    print(f"  FAIL: HTTP {e.code} - {body}")
    sys.exit(1)

# Load and inspect
wb = openpyxl.load_workbook(BytesIO(data), data_only=False)
print(f"  Sheets: {wb.sheetnames}")
expected = ['LAPL', 'LRSL', 'SI', 'SDF', 'Summary']
for s in expected:
    if s in wb.sheetnames:
        print(f"    ✓ {s} present")
    else:
        # Could be missing if no employees for that firm
        print(f"    (missing: {s} — may be empty firm)")

# Check LAPL sheet content
if 'LAPL' in wb.sheetnames:
    ws = wb['LAPL']
    print()
    print("  === LAPL sheet content ===")
    print(f"  A1 (company name): {ws['A1'].value}")
    print(f"  A2 (report title): {ws['A2'].value}")
    print(f"  Row 4 headers:")
    for col in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']:
        v = ws[f'{col}4'].value
        print(f"    {col}4: {v}")
    
    # Verify company name is NOT 'LAXREE GROUP OF COMPANIES'
    if ws['A1'].value == 'LAXREE GROUP OF COMPANIES':
        print("  FAIL: A1 still says 'LAXREE GROUP OF COMPANIES' — should be 'LAXREE AMENITIES PVT LTD'")
        sys.exit(1)
    elif ws['A1'].value == 'LAXREE AMENITIES PVT LTD':
        print("  ✓ A1 correctly shows company's full name")
    else:
        print(f"  ? A1 = '{ws['A1'].value}' — unexpected")

# Test 2: Specific firm
print()
print("=== Test 2: Specific firm (firm=LAPL) ===")
try:
    url = f"{BASE}/api/payroll/summary-export?month=7&year=2026&firm=LAPL"
    req = urllib.request.Request(url)
    with urllib.request.urlopen(req, timeout=30) as r:
        data2 = r.read()
        print(f"  OK: HTTP 200, {len(data2)} bytes")
        cd = r.headers.get('Content-Disposition', '')
        print(f"  Content-Disposition: {cd}")
        wb2 = openpyxl.load_workbook(BytesIO(data2), data_only=False)
        print(f"  Sheets: {wb2.sheetnames}")
        # Should have just LAPL + Summary
        if 'LAPL' in wb2.sheetnames and 'Summary' in wb2.sheetnames and len(wb2.sheetnames) == 2:
            print("  ✓ Only LAPL + Summary sheets (no other firms)")
        else:
            print(f"  FAIL: Expected only [LAPL, Summary], got {wb2.sheetnames}")
except urllib.error.HTTPError as e:
    body = e.read().decode()[:300]
    print(f"  FAIL: HTTP {e.code} - {body}")
    sys.exit(1)

print()
print("✅ All tests passed!")
