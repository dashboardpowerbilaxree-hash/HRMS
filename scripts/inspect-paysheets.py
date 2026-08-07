#!/usr/bin/env python3
"""Inspect the user's reference Payroll Summary file and our previously generated files."""
import openpyxl
from openpyxl.utils import get_column_letter

files = [
    "/home/z/my-project/upload/Payroll_Summary_July_2026_LAXREE GROUP OF COMPANIES.xlsx",
    "/home/z/my-project/upload/Payroll_July_2026.xlsx",
    "/home/z/my-project/upload/Payroll Master.xlsx",
]

for f in files:
    print(f"\n{'#'*100}")
    print(f"FILE: {f}")
    print(f"{'#'*100}")
    try:
        wb = openpyxl.load_workbook(f, data_only=False)
        print(f"Sheets: {wb.sheetnames}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"\n--- Sheet: {sn} (rows={ws.max_row}, cols={ws.max_column}) ---")
            for r in range(1, min(ws.max_row + 1, 60)):
                row_vals = []
                for c in range(1, min(ws.max_column + 1, 30)):
                    v = ws.cell(r, c).value
                    if v is not None:
                        row_vals.append(f"{get_column_letter(c)}{r}={v!r}")
                if row_vals:
                    print(f"  R{r}: " + " | ".join(row_vals))
    except Exception as e:
        print(f"ERROR: {e}")
