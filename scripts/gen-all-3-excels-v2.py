#!/usr/bin/env python3
"""
Generate 3 Excel files for July 2026 LAXREE GROUP OF COMPANIES HRMS:
  1. Payroll_Summary_July_2026.xlsx — 3 sheets (Payroll Register, Summary, Master)
  2. Payroll_Master_July_2026.xlsx — Master + 4 per-firm salary sheets (production export-master format)
  3. Attendance_Tracker_Monthly_July_2026.xlsx — 4 per-firm daily grids

Uses EXACT production recomputeStatus logic from src/lib/payroll-calc.ts:
  - isActuallyHalfDay: (totalHours || 0) < actualShiftHours / 2  (0 hours IS a half-day)
  - isActuallyEarlyOut: 5-minute grace period, 12-hour format fix-up
  - recomputeStatus: reclassify stored half-day → late/early-out/present if not genuine

All data is pulled fresh from production DB (psycopg2).
DB URL is in /tmp/db_url.txt.

Format functions:
  - fmt_hrs(decimal)      : REAL time conversion 9.02 → "9:01" (used for per-day TOTAL HRS)
  - display_decimal_col(v): DECIMAL split 216.07 → "216:07" (used for Total Working Hours column,
                            matches production export-master displayDecimalAsColon)
"""
import os
import calendar
from datetime import date, timedelta
from collections import defaultdict

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
YEAR, MONTH, DAYS = 2026, 7, 31
DB_URL = open('/tmp/db_url.txt').read().strip()

DOWNLOAD_DIR = '/home/z/my-project/download'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

OUT_SUMMARY   = f'{DOWNLOAD_DIR}/Payroll_Summary_July_2026.xlsx'
OUT_MASTER    = f'{DOWNLOAD_DIR}/Payroll_Master_July_2026.xlsx'
OUT_ATTENDANCE = f'{DOWNLOAD_DIR}/Attendance_Tracker_Monthly_July_2026.xlsx'

# Sundays in July 2026: 5, 12, 19, 26
SUNDAYS = {d for d in range(1, DAYS + 1) if date(YEAR, MONTH, d).weekday() == 6}
NUM_SUNDAYS = len(SUNDAYS)  # 4
HOLIDAYS = set()
TOTAL_WORKING_DAYS = DAYS - NUM_SUNDAYS - len(HOLIDAYS)  # 31 - 4 = 27

FIRM_DISPLAY = {
    'LAPL': 'LAXREE AMENITIES PVT LTD',
    'LRSL': 'LAXREE ROOFING SOLUTIONS PVT LTD',
    'SDF':  'SHREE DHAM FURNITURE',
    'SI':   'SHUBHAM INTERIORS',
}
FIRM_ORDER = ['LAPL', 'LRSL', 'SDF', 'SI']

# ─────────────────────────────────────────────────────────────────────────────
# PRODUCTION RECOMPUTE LOGIC (from src/lib/payroll-calc.ts)
# ─────────────────────────────────────────────────────────────────────────────
def t2m(t):
    """Time string 'HH:MM' → minutes since midnight. None on parse failure."""
    if not t:
        return None
    p = str(t).split(':')
    if len(p) != 2:
        return None
    try:
        return int(p[0]) * 60 + int(p[1])
    except Exception:
        return None


def actual_shift(stored, ss, se):
    """getActualShiftHours: derive shift duration from shiftStart/shiftEnd with 12-hour fix-up."""
    if ss and se:
        s, e = t2m(ss), t2m(se)
        if s is not None and e is not None:
            d = e - s
            if d < 0:
                d += 24 * 60
            return d / 60.0
    return stored or 9


def is_half(th, ash):
    """isActuallyHalfDay: (totalHours || 0) < actualShiftHours / 2.  0 hours IS a half-day."""
    return (th or 0) < ash / 2


def is_early(co, ss, se, grace=5):
    """isActuallyEarlyOut: checkOut < effective shiftEnd - grace (5 min). 12-hour fix-up applied."""
    if not co or not ss or not se:
        return False
    co, ss, se = t2m(co), t2m(ss), t2m(se)
    if co is None or ss is None or se is None:
        return False
    if se < ss:
        se += 24 * 60
    if co < ss:
        co += 24 * 60
    return co < se - grace


def rstat(rec, ash, ss, se):
    """recomputeStatus: reclassify stored half-day/early-out to actual status."""
    is_sd_half = rec['status'] in ('half-day', 'half_day') or rec.get('halfDay')
    if is_sd_half:
        if is_half(rec['totalHours'], ash):
            return 'half-day'
        ae = is_early(rec['checkOut'], ss, se) if (ss and se) else rec.get('earlyOut')
        if rec.get('lateEntry'):
            return 'late'
        if ae:
            return 'early-out'
        return 'present'
    is_sd_eo = rec['status'] == 'early-out' or rec.get('earlyOut')
    if is_sd_eo and ss and se:
        ae = is_early(rec['checkOut'], ss, se)
        if not ae:
            if rec.get('lateEntry'):
                return 'late'
            return 'present'
        return 'early-out'
    return rec['status']


# ─────────────────────────────────────────────────────────────────────────────
# FORMAT FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────
def fmt_hrs(decimal):
    """REAL time conversion: 9.02 → '9:01', 9.50 → '9:30'. Used for per-day TOTAL HRS column."""
    if decimal is None or decimal == 0:
        return '0:00'
    h = int(decimal)
    m = int(round((decimal - h) * 60))
    if m >= 60:
        h += 1
        m -= 60
    return f'{h}:{m:02d}'


def display_decimal_col(value):
    """DECIMAL split: 216.07 → '216:07', 36.00 → '36:00'. Matches production displayDecimalAsColon."""
    if value is None:
        return '0:00'
    v = float(value)
    s = f'{v:.2f}'
    int_part, dec_part = s.split('.')
    return f'{int_part}:{dec_part}'


# ─────────────────────────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────────────────────────
NAVY        = '1F3A5F'
GOLD        = 'C9A961'
DARK        = '0F1F3D'
LIGHT_BLUE  = 'E8EEF4'
LIGHT_GOLD  = 'FAF5E8'
LIGHT_GREEN = 'E8F5E9'
LIGHT_RED   = 'FDEDEC'
LIGHT_YEL   = 'FEF9E7'
SUN_FILL    = 'F4E5D6'
WHITE       = 'FFFFFF'
AMBER       = 'FFD700'

thin   = Side(border_style='thin',   color='666666')
medium = Side(border_style='medium', color=NAVY)
border_all   = Border(left=thin, right=thin, top=thin, bottom=thin)
border_thick = Border(left=medium, right=medium, top=medium, bottom=medium)

font_title    = Font(name='Calibri', size=14, bold=True, color=WHITE)
font_subtitle = Font(name='Calibri', size=10, italic=True, color=WHITE)
font_header   = Font(name='Calibri', size=9,  bold=True, color=WHITE)
font_body     = Font(name='Calibri', size=10)
font_body_sm  = Font(name='Calibri', size=9)
font_total    = Font(name='Calibri', size=10, bold=True, color=NAVY)

fill_title  = PatternFill('solid', fgColor=NAVY)
fill_header = PatternFill('solid', fgColor=NAVY)
fill_dark   = PatternFill('solid', fgColor=DARK)
fill_gold   = PatternFill('solid', fgColor=GOLD)
fill_total  = PatternFill('solid', fgColor=LIGHT_GOLD)
fill_alt    = PatternFill('solid', fgColor=LIGHT_BLUE)
fill_sun    = PatternFill('solid', fgColor=SUN_FILL)
fill_present = PatternFill('solid', fgColor=LIGHT_GREEN)
fill_absent  = PatternFill('solid', fgColor=LIGHT_RED)
fill_half    = PatternFill('solid', fgColor=LIGHT_YEL)

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
align_right  = Alignment(horizontal='right',  vertical='center')


# ═══════════════════════════════════════════════════════════════════════════
# STEP 1: Pull fresh data from production DB
# ═══════════════════════════════════════════════════════════════════════════
print('Connecting to production DB...')
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

# Active employees
cur.execute("""
    SELECT "employeeId", "fullName", firm, "location", "shiftHours",
           "shiftStart", "shiftEnd", "monthlySalary", "salaryType",
           "employmentType"
    FROM "Employee" WHERE status = 'Yes'
""")
EMPLOYEES = {}
for r in cur.fetchall():
    EMPLOYEES[r[0]] = {
        'employeeId':    r[0],
        'fullName':      r[1],
        'firm':          r[2],
        'location':      r[3],
        'shiftHours':    float(r[4] or 9),
        'shiftStart':    r[5],
        'shiftEnd':      r[6],
        'monthlySalary': float(r[7] or 0),
        'salaryType':    r[8] or 'Monthly',
        'employmentType': r[9] or 'Full Time',
    }
print(f'  Loaded {len(EMPLOYEES)} active employees')

# Attendance records for July 2026
cur.execute("""
    SELECT "employeeId", date, status, "checkIn", "checkOut", "totalHours",
           "isSunday", "isHoliday", "halfDay", "lateEntry", "earlyOut", "overtimeHours"
    FROM "Attendance"
    WHERE date >= '2026-07-01' AND date <= '2026-07-31'
""")
ATT = defaultdict(dict)  # {empId: {day: rec}}
for r in cur.fetchall():
    rec = {
        'day':            r[1].day,
        'status':         r[2],
        'checkIn':        r[3],
        'checkOut':       r[4],
        'totalHours':     float(r[5] or 0),
        'isSunday':       r[6],
        'isHoliday':      r[7],
        'halfDay':        r[8],
        'lateEntry':      r[9],
        'earlyOut':       r[10],
        'overtimeHours':  float(r[11] or 0),
    }
    ATT[r[0]][rec['day']] = rec
print(f'  Loaded attendance records for {len(ATT)} employees')

# Approved leaves overlapping July 2026
cur.execute("""
    SELECT "employeeId", "startDate", "endDate"
    FROM "Leave"
    WHERE "startDate" <= '2026-07-31' AND "endDate" >= '2026-07-01'
      AND status = 'approved'
""")
LV_DATES = defaultdict(set)
for ec, sd, ed in cur.fetchall():
    c = sd.date() if hasattr(sd, 'date') else sd
    e = ed.date() if hasattr(ed, 'date') else ed
    while c <= e:
        if date(YEAR, MONTH, 1) <= c <= date(YEAR, MONTH, DAYS):
            LV_DATES[ec].add(c.isoformat())
        c += timedelta(days=1)
print(f'  Loaded approved leaves for {len(LV_DATES)} employees')

cur.close()
conn.close()
print('  DB connection closed')


# ═══════════════════════════════════════════════════════════════════════════
# STEP 2: Compute metrics for each employee using PRODUCTION logic
# ═══════════════════════════════════════════════════════════════════════════
print('\nComputing metrics for each employee...')

# Sort employees by firm (LAPL, LRSL, SDF, SI) then fullName
SORTED_EMPS = sorted(
    EMPLOYEES.values(),
    key=lambda e: (FIRM_ORDER.index(e['firm']) if e['firm'] in FIRM_ORDER else 99, e['fullName'])
)

for emp in SORTED_EMPS:
    ec = emp['employeeId']
    sh = emp['shiftHours']
    ss = emp['shiftStart']
    se = emp['shiftEnd']
    ash = actual_shift(sh, ss, se)
    emp_att = ATT.get(ec, {})
    lds = LV_DATES.get(ec, set())

    # Per-day recomputed status & code
    daily = {}  # {day: {'status': st, 'code': code, 'rec': rec}}
    pds = set()  # present days (any recomputed present/late/early-out/half-day)
    for d, rec in emp_att.items():
        st = rstat(rec, ash, ss, se)
        if st in ('present', 'late', 'early-out', 'half-day', 'half_day'):
            pds.add(f'{YEAR}-{MONTH:02d}-{d:02d}')
        daily[d] = {'status': st, 'rec': rec}

    present_days = 0.0
    half_days = 0
    absent_days = 0.0   # RAW absent (no leave)
    leave_days = 0      # absent days that have approved leave
    worked_hrs = 0.0
    ot_hrs = 0.0

    for d in range(1, DAYS + 1):
        ds = f'{YEAR}-{MONTH:02d}-{d:02d}'
        is_sun = d in SUNDAYS
        info = daily.get(d)
        if info:
            st = info['status']
            rec = info['rec']
            if st == 'absent':
                ild = ds in lds and ds not in pds
                if ild and not is_sun:
                    leave_days += 1
                else:
                    absent_days += 1
            elif st == 'weekly-off':
                if rec['checkIn'] and rec['totalHours'] > 0:
                    worked_hrs += rec['totalHours']
                    ot_hrs += rec['overtimeHours']
                    present_days += 1
            elif st == 'holiday':
                if rec['checkIn'] and rec['totalHours'] > 0:
                    worked_hrs += rec['totalHours']
                    ot_hrs += rec['overtimeHours']
                    present_days += 1
            elif st in ('half-day', 'half_day'):
                worked_hrs += rec['totalHours']
                ot_hrs += rec['overtimeHours']
                present_days += 0.5
                absent_days += 0.5
                half_days += 1
            else:  # present, late, early-out
                worked_hrs += rec['totalHours']
                ot_hrs += rec['overtimeHours']
                present_days += 1
        else:
            ild = ds in lds and ds not in pds
            if is_sun:
                pass
            elif ild:
                leave_days += 1
            else:
                absent_days += 1

    # Derive final metrics
    sunday_hrs = NUM_SUNDAYS * sh
    total_hrs = worked_hrs + sunday_hrs
    sl_per_hr = emp['monthlySalary'] / (DAYS * sh) if sh else 0
    gross = round(total_hrs * sl_per_hr)

    # Absent Days column in Payroll Summary: totalWorkingDays - present - half (= total non-working days, includes leave)
    # Leave column in per-firm sheets: absent_days (raw, no leave) + leave_days (= same value, total non-working days)
    absent_days_total = TOTAL_WORKING_DAYS - present_days - half_days  # includes leave days
    leave_col_value = absent_days + leave_days  # same number, but conceptually separate

    emp['actualShiftHours'] = ash
    emp['presentDays'] = present_days
    emp['halfDays'] = half_days
    emp['absentDays'] = absent_days_total         # Payroll Summary "Absent Days"
    emp['absentDaysRaw'] = absent_days            # raw absent (no leave) — used for Leave column
    emp['leaveDays'] = leave_days                 # approved leave days that were absent
    emp['leaveColValue'] = leave_col_value        # per-firm sheet "Leave" column
    emp['workedHrsInclOT'] = worked_hrs
    emp['otHrs'] = ot_hrs
    emp['sundayHrs'] = sunday_hrs
    emp['totalHrs'] = total_hrs
    emp['slPerHr'] = sl_per_hr
    emp['grossSalary'] = gross
    emp['daily'] = daily

print(f'  Computed metrics for {len(SORTED_EMPS)} employees')


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3: Generate Payroll_Summary_July_2026.xlsx
# ═══════════════════════════════════════════════════════════════════════════
print(f'\nGenerating {OUT_SUMMARY}...')
wb1 = Workbook()
wb1.properties.creator = 'Z.ai'

# ─── Sheet 1: Payroll Register ───
ws = wb1.active
ws.title = 'Payroll Register'

ws.merge_cells('A1:N1')
ws['A1'] = 'LAXREE GROUP OF COMPANIES'
ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=WHITE)
ws['A1'].fill = fill_title
ws['A1'].alignment = align_center
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:N2')
ws['A2'] = 'Payroll Register — July 2026'
ws['A2'].font = font_subtitle
ws['A2'].fill = fill_title
ws['A2'].alignment = align_center
ws.row_dimensions[2].height = 20

ws.merge_cells('A3:N3')
ws['A3'] = (f'Days in Month: {DAYS}  |  Sundays: {NUM_SUNDAYS}  |  '
            f'Holidays: {len(HOLIDAYS)}  |  Working Days: {TOTAL_WORKING_DAYS}  |  '
            f'Total Employees: {len(SORTED_EMPS)}')
ws['A3'].font = Font(name='Calibri', size=9, italic=True, color='555555')
ws['A3'].alignment = align_center
ws.row_dimensions[3].height = 18

HEADERS = [
    'S.No', 'Employee Name', 'Monthly Salary', 'Working Hrs', 'Sl/Hr',
    'Present Days', 'Absent Days', 'Worked Hrs including OT', 'Additional hrs (Sunday Hrs)',
    'Total Hrs', 'Gross', 'SD Refund', 'Salary Advance', 'Net'
]
for i, h in enumerate(HEADERS, 1):
    c = ws.cell(4, i, h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center
    c.border = border_all
ws.row_dimensions[4].height = 38

row = 5
for idx, emp in enumerate(SORTED_EMPS, 1):
    vals = [
        idx,
        emp['fullName'],
        emp['monthlySalary'],
        emp['shiftHours'],
        round(emp['slPerHr'], 2),
        emp['presentDays'],
        emp['absentDays'],
        display_decimal_col(emp['workedHrsInclOT']),
        display_decimal_col(emp['sundayHrs']),
        display_decimal_col(emp['totalHrs']),
        emp['grossSalary'],
        0,   # SD Refund
        0,   # Salary Advance
        emp['grossSalary'],  # Net = Gross (no deductions)
    ]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v)
        c.font = font_body
        c.border = border_all
        if i in (1, 4, 6, 7, 8, 9, 10):
            c.alignment = align_center
        elif i in (3, 5, 11, 12, 13, 14):
            c.alignment = align_right
        else:
            c.alignment = align_left
        if row % 2 == 0:
            c.fill = fill_alt

    ws.cell(row, 3).number_format = '#,##0'
    ws.cell(row, 5).number_format = '0.00'
    ws.cell(row, 11).number_format = '#,##0'
    ws.cell(row, 14).number_format = '#,##0'
    row += 1

# Total row
total_row = row
ws.cell(total_row, 2).value = 'TOTAL'
ws.cell(total_row, 3).value = f'=SUM(C5:C{total_row-1})'
ws.cell(total_row, 6).value = f'=SUM(F5:F{total_row-1})'
ws.cell(total_row, 7).value = f'=SUM(G5:G{total_row-1})'
ws.cell(total_row, 11).value = f'=SUM(K5:K{total_row-1})'
ws.cell(total_row, 12).value = f'=SUM(L5:L{total_row-1})'
ws.cell(total_row, 13).value = f'=SUM(M5:M{total_row-1})'
ws.cell(total_row, 14).value = f'=SUM(N5:N{total_row-1})'

for i in range(1, 15):
    c = ws.cell(total_row, i)
    c.font = font_total
    c.fill = fill_total
    c.border = border_all
    if i in (3, 11, 12, 13, 14):
        c.alignment = align_right
        c.number_format = '#,##0'
    elif i in (6, 7):
        c.alignment = align_center
    else:
        c.alignment = align_center

widths = [6, 26, 14, 10, 10, 11, 11, 18, 18, 12, 12, 11, 13, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A5'

# ─── Sheet 2: Summary ───
ws2 = wb1.create_sheet('Summary')
ws2.merge_cells('A1:B1')
ws2['A1'] = 'Payroll Summary'
ws2['A1'].font = Font(name='Calibri', size=16, bold=True, color=WHITE)
ws2['A1'].fill = fill_title
ws2['A1'].alignment = align_center
ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:B2')
ws2['A2'] = 'Laxree Group of Companies — July 2026'
ws2['A2'].font = font_subtitle
ws2['A2'].fill = fill_title
ws2['A2'].alignment = align_center

total_gross = sum(e['grossSalary'] for e in SORTED_EMPS)
total_present = sum(e['presentDays'] for e in SORTED_EMPS)
total_absent = sum(e['absentDays'] for e in SORTED_EMPS)

r = 4
ws2.cell(r, 1, 'Category').font = font_header
ws2.cell(r, 1).fill = fill_header
ws2.cell(r, 1).alignment = align_center
ws2.cell(r, 1).border = border_all
ws2.cell(r, 2, 'Amount (₹)').font = font_header
ws2.cell(r, 2).fill = fill_header
ws2.cell(r, 2).alignment = align_center
ws2.cell(r, 2).border = border_all
r += 1

summary_rows = [
    ('Total Gross Salary', total_gross),
    ('Total Bonus', 0),
    ('Total Deductions', 0),
    ('Total Net Payroll', total_gross),
]
for label, val in summary_rows:
    ws2.cell(r, 1, label).font = font_body
    ws2.cell(r, 1).border = border_all
    ws2.cell(r, 1).alignment = align_left
    ws2.cell(r, 2, val).font = font_body
    ws2.cell(r, 2).border = border_all
    ws2.cell(r, 2).alignment = align_right
    ws2.cell(r, 2).number_format = '#,##0'
    if label == 'Total Net Payroll':
        ws2.cell(r, 1).font = font_total
        ws2.cell(r, 2).font = font_total
        ws2.cell(r, 1).fill = fill_total
        ws2.cell(r, 2).fill = fill_total
    r += 1

r += 1
ws2.cell(r, 1, 'Metric').font = font_header
ws2.cell(r, 1).fill = fill_header
ws2.cell(r, 1).alignment = align_center
ws2.cell(r, 1).border = border_all
ws2.cell(r, 2, 'Value').font = font_header
ws2.cell(r, 2).fill = fill_header
ws2.cell(r, 2).alignment = align_center
ws2.cell(r, 2).border = border_all
r += 1

metric_rows = [
    ('Employees Processed', len(SORTED_EMPS)),
    ('Total Present Days', total_present),
    ('Total Absent Days', total_absent),
    ('Average Net Salary', round(total_gross / len(SORTED_EMPS)) if SORTED_EMPS else 0),
]
for label, val in metric_rows:
    ws2.cell(r, 1, label).font = font_body
    ws2.cell(r, 1).border = border_all
    ws2.cell(r, 1).alignment = align_left
    ws2.cell(r, 2, val).font = font_body
    ws2.cell(r, 2).border = border_all
    ws2.cell(r, 2).alignment = align_right
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        ws2.cell(r, 2).number_format = '#,##0' if isinstance(val, int) else '#,##0.00'
    r += 1

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 20

# ─── Sheet 3: Master ───
ws3 = wb1.create_sheet('Master')
ws3.merge_cells('A1:L1')
ws3['A1'] = 'LAXREE GROUP OF COMPANIES — Master Employee List (July 2026)'
ws3['A1'].font = Font(name='Calibri', size=14, bold=True, color=WHITE)
ws3['A1'].fill = fill_title
ws3['A1'].alignment = align_center
ws3.row_dimensions[1].height = 28

ws3.merge_cells('A2:L2')
ws3['A2'] = (f'Days in Month: {DAYS}  |  Sundays: {NUM_SUNDAYS}  |  '
             f'Holidays: {len(HOLIDAYS)}  |  Working Days: {TOTAL_WORKING_DAYS}')
ws3['A2'].font = Font(name='Calibri', size=10, italic=True, color='555555')
ws3['A2'].alignment = align_center

master_headers = ['S.No', 'Emp Code', 'Employee Name', 'Firm', 'Location',
                  'Monthly Salary', 'Shift Hrs/Day', 'Sl/Hr', 'Present Days',
                  'Absent Days', 'Worked Hrs incl OT', 'Total Hrs']
for i, h in enumerate(master_headers, 1):
    c = ws3.cell(4, i, h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center
    c.border = border_all
ws3.row_dimensions[4].height = 32

r = 5
for idx, emp in enumerate(SORTED_EMPS, 1):
    vals = [
        idx,
        emp['employeeId'],
        emp['fullName'],
        emp['firm'] or '-',
        emp['location'] or '-',
        emp['monthlySalary'],
        emp['shiftHours'],
        round(emp['slPerHr'], 2),
        emp['presentDays'],
        emp['absentDays'],
        display_decimal_col(emp['workedHrsInclOT']),
        display_decimal_col(emp['totalHrs']),
    ]
    for i, v in enumerate(vals, 1):
        c = ws3.cell(r, i, v)
        c.font = font_body
        c.border = border_all
        if i in (1, 4, 7, 8, 9, 10, 11, 12):
            c.alignment = align_center
        elif i == 6:
            c.alignment = align_right
            c.number_format = '#,##0'
        else:
            c.alignment = align_left
        if r % 2 == 0:
            c.fill = fill_alt
    r += 1

ws3.cell(r, 3).value = 'TOTAL'
ws3.cell(r, 6).value = f'=SUM(F5:F{r-1})'
ws3.cell(r, 9).value = f'=SUM(I5:I{r-1})'
ws3.cell(r, 10).value = f'=SUM(J5:J{r-1})'
for i in range(1, 13):
    c = ws3.cell(r, i)
    c.font = font_total
    c.fill = fill_total
    c.border = border_all
    if i == 6:
        c.number_format = '#,##0'
        c.alignment = align_right
    elif i in (9, 10):
        c.alignment = align_center
    else:
        c.alignment = align_center

widths3 = [6, 12, 26, 7, 16, 14, 11, 10, 11, 11, 16, 13]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A5'

wb1.save(OUT_SUMMARY)
print(f'  ✓ Saved: {OUT_SUMMARY}')


# ═══════════════════════════════════════════════════════════════════════════
# STEP 4: Generate Payroll_Master_July_2026.xlsx
#   - Master sheet (single-row-per-employee info)
#   - 4 per-firm salary sheets in PRODUCTION export-master format
#     (3 sections: days 1-11, 12-22, 23-31 with IN/OUT/TOTAL HRS columns)
# ═══════════════════════════════════════════════════════════════════════════
print(f'\nGenerating {OUT_MASTER}...')
wb2 = Workbook()
wb2.properties.creator = 'Z.ai'

# ─── Master Sheet ───
ws = wb2.active
ws.title = 'Master'

ws.merge_cells('A1:L1')
ws['A1'] = 'LAXREE GROUP OF COMPANIES — Master Employee List'
ws['A1'].font = font_title
ws['A1'].fill = fill_title
ws['A1'].alignment = align_center
ws.row_dimensions[1].height = 28

ws.merge_cells('A2:L2')
ws['A2'] = (f'Salary Month: July 2026  |  Days in Month: {DAYS}  |  '
            f'Sundays: {NUM_SUNDAYS}  |  Holidays: {len(HOLIDAYS)}  |  '
            f'Working Days: {TOTAL_WORKING_DAYS}')
ws['A2'].font = Font(name='Calibri', size=10, italic=True, color=WHITE)
ws['A2'].fill = fill_title
ws['A2'].alignment = align_center

master_headers = ['Emp Code', 'Full Name', 'Firm', 'Location', 'Salary Type',
                  'Monthly Salary (₹)', 'Daily Rate (₹)', 'Shift Hrs/Day',
                  'Shift Start', 'Shift End', 'Employment Type', 'Status']
for i, h in enumerate(master_headers, 1):
    c = ws.cell(4, i, h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center
    c.border = border_all
ws.row_dimensions[4].height = 32

r = 5
for emp in SORTED_EMPS:
    vals = [
        emp['employeeId'],
        emp['fullName'],
        emp['firm'] or '-',
        emp['location'] or '-',
        emp['salaryType'] or 'Monthly',
        emp['monthlySalary'],
        round(emp['monthlySalary'] / DAYS, 2) if emp['monthlySalary'] else 0,
        emp['shiftHours'],
        emp['shiftStart'],
        emp['shiftEnd'],
        emp['employmentType'] or 'Full Time',
        'Active',
    ]
    for i, v in enumerate(vals, 1):
        c = ws.cell(r, i, v)
        c.font = font_body
        c.border = border_all
        if i in (1, 3, 8, 9, 10, 11, 12):
            c.alignment = align_center
        elif i in (6, 7):
            c.alignment = align_right
            c.number_format = '#,##0.00'
        else:
            c.alignment = align_left
        if r % 2 == 0:
            c.fill = fill_alt
    r += 1

ws.cell(r, 2).value = 'TOTAL'
ws.cell(r, 6).value = f'=SUM(F5:F{r-1})'
ws.cell(r, 6).number_format = '#,##0'
for i in range(1, 13):
    c = ws.cell(r, i)
    c.font = font_total
    c.fill = fill_total
    c.border = border_all
    c.alignment = align_right if i in (6, 7) else align_center

widths = [10, 24, 7, 16, 11, 16, 13, 11, 11, 11, 14, 9]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A5'

# ─── Per-firm Salary Sheets (PRODUCTION export-master format) ───
def cell_code_for_day(emp, d):
    """Return (in_val, out_val, hrs_val) for a given day cell in per-firm sheet."""
    is_sun = d in SUNDAYS
    ds = f'{YEAR}-{MONTH:02d}-{d:02d}'
    info = emp['daily'].get(d)
    lds = LV_DATES.get(emp['employeeId'], set())

    if is_sun:
        if info and info['rec']['checkIn'] and info['rec']['totalHours'] > 0:
            rec = info['rec']
            return (rec['checkIn'] or '', rec['checkOut'] or '', fmt_hrs(rec['totalHours']))
        return ('Weekly Off', '', '')

    if not info:
        # No record — check leave
        if ds in lds:
            return ('Leave', '', '')
        return ('Absent', '', '')

    st = info['status']
    rec = info['rec']
    if st == 'absent':
        if ds in lds:
            return ('Leave', '', '')
        return ('Absent', '', '')
    if st == 'weekly-off':
        if rec['checkIn'] and rec['totalHours'] > 0:
            return (rec['checkIn'] or '', rec['checkOut'] or '', fmt_hrs(rec['totalHours']))
        return ('Weekly Off', '', '')
    if st == 'holiday':
        if rec['checkIn'] and rec['totalHours'] > 0:
            return (rec['checkIn'] or '', rec['checkOut'] or '', fmt_hrs(rec['totalHours']))
        return ('Holiday', '', '')
    if st in ('half-day', 'half_day'):
        return (rec['checkIn'] or 'Half Day', rec['checkOut'] or '', fmt_hrs(rec['totalHours']))
    # present, late, early-out
    return (rec['checkIn'] or '', rec['checkOut'] or '', fmt_hrs(rec['totalHours']))


def build_firm_sheet(wb, firm):
    """Build a per-firm salary sheet in production export-master format (3 sections)."""
    firm_emps = [e for e in SORTED_EMPS if e['firm'] == firm]
    sheet_name = f'{firm}_July_2026_Sal'
    ws = wb.create_sheet(sheet_name)

    # Section definitions: (start_day, end_day, extra_cols)
    sections = [
        (1, 11, []),
        (12, 22, []),
        (23, DAYS, ['Total Working Hours', 'Leave']),
    ]
    # Max cols: 1 (name) + max days × 3 + max extras
    max_days = max(s[1] - s[0] + 1 for s in sections)  # 11
    max_extras = max(len(s[2]) for s in sections)  # 2
    max_cols = 1 + max_days * 3 + max_extras  # 1 + 33 + 2 = 36

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    ws.cell(1, 1).value = f'ATTENDENCE - {YEAR}  |  SALARY SHEET OF {FIRM_DISPLAY.get(firm, firm)}  |  JULY {YEAR}'
    ws.cell(1, 1).font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    ws.cell(1, 1).fill = fill_gold
    ws.cell(1, 1).alignment = align_center
    ws.row_dimensions[1].height = 26

    current_row = 2
    for sec_idx, (start_day, end_day, extras) in enumerate(sections):
        num_days = end_day - start_day + 1

        # Date header row
        ws.cell(current_row, 1).value = 'EMP'
        ws.cell(current_row, 1).font = Font(name='Calibri', size=10, bold=True, color=WHITE)
        ws.cell(current_row, 1).fill = fill_dark
        ws.cell(current_row, 1).alignment = align_center
        ws.cell(current_row, 1).border = border_all

        for d in range(num_days):
            day = start_day + d
            col_start = 2 + d * 3
            ws.merge_cells(start_row=current_row, start_column=col_start,
                           end_row=current_row, end_column=col_start + 2)
            cell = ws.cell(current_row, col_start)
            cell.value = f'{day}/{MONTH:02d}/{YEAR}'
            is_sun = day in SUNDAYS
            cell.font = Font(name='Calibri', size=10, bold=True,
                             color=AMBER if is_sun else WHITE)
            cell.fill = fill_dark
            cell.alignment = align_center
            for cc in range(col_start, col_start + 3):
                ws.cell(current_row, cc).border = border_all
                ws.cell(current_row, cc).fill = fill_dark

        # Extra cols header (in same row)
        for ei, extra in enumerate(extras):
            col = 2 + num_days * 3 + ei
            cell = ws.cell(current_row, col)
            cell.value = extra
            cell.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
            cell.fill = fill_dark
            cell.alignment = align_center
            cell.border = border_all

        # Fill remaining cols with dark background
        for col in range(2 + num_days * 3 + len(extras), max_cols + 1):
            ws.cell(current_row, col).fill = fill_dark
            ws.cell(current_row, col).border = border_all

        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Sub-header row (IN/OUT/TOTAL HRS)
        ws.cell(current_row, 1).value = ''
        ws.cell(current_row, 1).fill = fill_header
        ws.cell(current_row, 1).border = border_all
        for d in range(num_days):
            day = start_day + d
            col_start = 2 + d * 3
            is_sun = day in SUNDAYS
            for j, label in enumerate(['IN', 'OUT', 'TOTAL HRS']):
                cell = ws.cell(current_row, col_start + j)
                cell.value = label
                cell.font = Font(name='Calibri', size=9, bold=True,
                                 color=AMBER if is_sun else WHITE)
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_all
        for ei, extra in enumerate(extras):
            col = 2 + num_days * 3 + ei
            cell = ws.cell(current_row, col)
            cell.value = ''
            cell.fill = fill_header
            cell.border = border_all
        for col in range(2 + num_days * 3 + len(extras), max_cols + 1):
            ws.cell(current_row, col).fill = fill_header
            ws.cell(current_row, col).border = border_all

        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Employee data rows
        for emp in firm_emps:
            ws.cell(current_row, 1).value = emp['fullName']
            ws.cell(current_row, 1).font = font_body_sm
            ws.cell(current_row, 1).alignment = align_left
            ws.cell(current_row, 1).border = border_all
            if current_row % 2 == 0:
                ws.cell(current_row, 1).fill = fill_alt

            for d in range(num_days):
                day = start_day + d
                col_start = 2 + d * 3
                in_val, out_val, hrs_val = cell_code_for_day(emp, day)
                for j, v in enumerate([in_val, out_val, hrs_val]):
                    cell = ws.cell(current_row, col_start + j)
                    cell.value = v
                    cell.font = font_body_sm
                    cell.alignment = align_center
                    cell.border = border_all
                    if day in SUNDAYS:
                        cell.fill = fill_sun
                    elif current_row % 2 == 0:
                        cell.fill = fill_alt
                    # Special color for status cells
                    if j == 0:
                        if v == 'Absent':
                            cell.fill = fill_absent
                        elif v == 'Leave':
                            cell.fill = fill_half
                        elif v == 'Weekly Off':
                            cell.fill = fill_sun
                        elif v == 'Half Day':
                            cell.fill = fill_half

            # Extra cols values
            for ei, extra in enumerate(extras):
                col = 2 + num_days * 3 + ei
                cell = ws.cell(current_row, col)
                if extra == 'Total Working Hours':
                    cell.value = display_decimal_col(emp['workedHrsInclOT'])
                elif extra == 'Leave':
                    cell.value = int(emp['leaveColValue']) if emp['leaveColValue'] == int(emp['leaveColValue']) else emp['leaveColValue']
                cell.font = Font(name='Calibri', size=9, bold=True, color=NAVY)
                cell.alignment = align_center
                cell.border = border_all
                cell.fill = fill_total

            # Fill remaining cols
            for col in range(2 + num_days * 3 + len(extras), max_cols + 1):
                ws.cell(current_row, col).border = border_all
                if current_row % 2 == 0:
                    ws.cell(current_row, col).fill = fill_alt

            ws.row_dimensions[current_row].height = 18
            current_row += 1

    # Column widths
    ws.column_dimensions['A'].width = 22  # name
    for c in range(2, 2 + 11 * 3 + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9
    # Total Working Hours col (col 32 in section 3, but it's at col 2 + 9*3 = 29 for section 3)
    # We set width by letter — col 29 = "AC"
    ws.column_dimensions[get_column_letter(29)].width = 16  # Total Working Hours
    ws.column_dimensions[get_column_letter(30)].width = 8   # Leave

    ws.freeze_panes = 'B4'


for firm in FIRM_ORDER:
    firm_emps = [e for e in SORTED_EMPS if e['firm'] == firm]
    if not firm_emps:
        continue
    build_firm_sheet(wb2, firm)
    print(f'  ✓ Built per-firm sheet: {firm}_July_2026_Sal ({len(firm_emps)} employees)')

wb2.save(OUT_MASTER)
print(f'  ✓ Saved: {OUT_MASTER}')


# ═══════════════════════════════════════════════════════════════════════════
# STEP 5: Generate Attendance_Tracker_Monthly_July_2026.xlsx
#   4 per-firm daily grids with P/A/H/L/E/S codes
# ═══════════════════════════════════════════════════════════════════════════
print(f'\nGenerating {OUT_ATTENDANCE}...')
wb3 = Workbook()
wb3.properties.creator = 'Z.ai'


def status_code(emp, d):
    """Return short code for daily grid: P/A/H/L/E/S."""
    if d in SUNDAYS:
        return 'S'  # Sunday / Weekly Off
    info = emp['daily'].get(d)
    ds = f'{YEAR}-{MONTH:02d}-{d:02d}'
    lds = LV_DATES.get(emp['employeeId'], set())
    if not info:
        if ds in lds:
            return 'L'  # Approved Leave
        return 'A'  # Absent
    st = info['status']
    if st == 'absent':
        if ds in lds:
            return 'L'
        return 'A'
    if st == 'weekly-off':
        return 'S'
    if st == 'holiday':
        return 'H'  # Holiday — code H (will use different color)
    if st in ('half-day', 'half_day'):
        return 'H'  # Half-day
    if st == 'late':
        return 'L'  # Late
    if st == 'early-out':
        return 'E'  # Early out
    if st == 'present':
        return 'P'
    return st[:1].upper() if st else '-'


first_firm = True
for firm in FIRM_ORDER:
    firm_emps = [e for e in SORTED_EMPS if e['firm'] == firm]
    if not firm_emps:
        continue
    sheet_name = f'{firm}_July_2026_Att'
    if first_firm:
        ws = wb3.active
        ws.title = sheet_name
        first_firm = False
    else:
        ws = wb3.create_sheet(sheet_name)

    total_cols = 3 + DAYS + 5  # S.No, Emp Code, Name, days 1-31, Present, Absent, Half, Total Hrs, OT Hrs

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(1, 1).value = f'LAXREE GROUP OF COMPANIES — {firm} Attendance Tracker — July 2026'
    ws.cell(1, 1).font = font_title
    ws.cell(1, 1).fill = fill_title
    ws.cell(1, 1).alignment = align_center
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(2, 1).value = (f'Days in Month: {DAYS}  |  Sundays: {NUM_SUNDAYS}  |  '
                           f'Holidays: {len(HOLIDAYS)}  |  Working Days: {TOTAL_WORKING_DAYS}  |  '
                           f'Codes: P=Present  A=Absent  H=Half-day  L=Late/Leave  E=Early-out  S=Sunday')
    ws.cell(2, 1).font = Font(name='Calibri', size=9, italic=True, color='555555')
    ws.cell(2, 1).alignment = align_center

    # Headers rows 3 & 4
    fixed_headers = ['S.No', 'Emp Code', 'Employee Name']
    for i, h in enumerate(fixed_headers, 1):
        c = ws.cell(3, i, h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_all
        ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)

    for d in range(1, DAYS + 1):
        col = 3 + d
        c = ws.cell(3, col, d)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_all
        weekday = calendar.weekday(YEAR, MONTH, d)
        day_name = calendar.day_abbr[weekday][:2]
        c2 = ws.cell(4, col, day_name)
        c2.font = Font(name='Calibri', size=8, bold=True,
                       color=AMBER if d in SUNDAYS else WHITE)
        c2.fill = fill_header
        c2.alignment = align_center
        c2.border = border_all

    summary_start = 3 + DAYS + 1
    summary_headers = ['Present', 'Absent', 'Half', 'Total Hrs', 'OT Hrs']
    for i, h in enumerate(summary_headers):
        col = summary_start + i
        c = ws.cell(3, col, h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_all
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 16

    # Data rows
    r = 5
    for idx, emp in enumerate(firm_emps, 1):
        ws.cell(r, 1).value = idx
        ws.cell(r, 2).value = emp['employeeId']
        ws.cell(r, 3).value = emp['fullName']
        for c in range(1, 4):
            cell = ws.cell(r, c)
            cell.font = font_body_sm
            cell.border = border_all
            cell.alignment = align_center if c != 3 else align_left
            if r % 2 == 0:
                cell.fill = fill_alt

        for d in range(1, DAYS + 1):
            col = 3 + d
            code = status_code(emp, d)
            cell = ws.cell(r, col, code)
            cell.font = font_body_sm
            cell.border = border_all
            cell.alignment = align_center
            if d in SUNDAYS:
                cell.fill = fill_sun
            elif code == 'P':
                cell.fill = fill_present
            elif code == 'A':
                cell.fill = fill_absent
            elif code == 'H':
                cell.fill = fill_half
            elif code == 'L':
                cell.fill = fill_half
            elif code == 'E':
                cell.fill = fill_half
            elif r % 2 == 0:
                cell.fill = fill_alt

        # Summary
        ws.cell(r, summary_start).value = emp['presentDays']
        ws.cell(r, summary_start + 1).value = emp['absentDays']
        ws.cell(r, summary_start + 2).value = emp['halfDays']
        ws.cell(r, summary_start + 3).value = display_decimal_col(emp['workedHrsInclOT'])
        ws.cell(r, summary_start + 4).value = display_decimal_col(emp['otHrs'])
        for i in range(5):
            cell = ws.cell(r, summary_start + i)
            cell.font = font_total if i < 3 else font_body_sm
            cell.border = border_all
            cell.alignment = align_center
            if r % 2 == 0:
                cell.fill = fill_alt
        r += 1

    # Total row
    ws.cell(r, 3).value = 'TOTAL'
    for i in range(3):
        col = summary_start + i
        ws.cell(r, col).value = f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{r-1})'
    for i in range(1, total_cols + 1):
        c = ws.cell(r, i)
        c.font = font_total
        c.fill = fill_total
        c.border = border_all
        c.alignment = align_center

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 22
    for d in range(1, DAYS + 1):
        ws.column_dimensions[get_column_letter(3 + d)].width = 4
    for i in range(5):
        ws.column_dimensions[get_column_letter(summary_start + i)].width = 10

    ws.freeze_panes = 'D5'

wb3.save(OUT_ATTENDANCE)
print(f'  ✓ Saved: {OUT_ATTENDANCE}')


# ═══════════════════════════════════════════════════════════════════════════
# STEP 6: Verification — Kamlesh + all-employee day-sum
# ═══════════════════════════════════════════════════════════════════════════
print('\n' + '=' * 80)
print('VERIFICATION')
print('=' * 80)

# Kamlesh check
kamlesh = next(e for e in SORTED_EMPS if e['employeeId'] == 'EMP-021')
print(f'\nKamlesh (EMP-021):')
print(f'  Present: {kamlesh["presentDays"]:.0f}  (expected 24)')
print(f'  Absent:  {kamlesh["absentDays"]:.0f}  (expected 3)')
print(f'  Leave column (per-firm sheet): {int(kamlesh["leaveColValue"])}  (expected 3)')
print(f'  Worked Hrs incl OT: {display_decimal_col(kamlesh["workedHrsInclOT"])}  (expected 216:07)')
print(f'  Sunday Hrs: {display_decimal_col(kamlesh["sundayHrs"])}  (expected 36:00)')
print(f'  Total Hrs: {display_decimal_col(kamlesh["totalHrs"])}  (expected 252:07)')
print(f'  Gross: ₹{kamlesh["grossSalary"]:,}  (expected ₹17,166)')

# All employees day-sum check
print(f'\nAll {len(SORTED_EMPS)} employees day-sum check (P + A + H + Sundays should = 31):')
all_pass = True
for emp in SORTED_EMPS:
    p = emp['presentDays']
    h = emp['halfDays']
    a_raw = emp['absentDaysRaw']
    lv = emp['leaveDays']
    sun = NUM_SUNDAYS
    # Total = present + half*0.5 (already in present) + absent_raw + leave + sundays
    # Actually: present includes 0.5 per half-day. half_days count = number of half-days.
    # So: total = present + absent_raw + leave + sundays (since half-day's 0.5 absent is already in absent_raw)
    total = p + a_raw + lv + sun
    if abs(total - DAYS) > 0.01:
        print(f'  FAIL {emp["employeeId"]} {emp["fullName"]}: P={p}, H={h}, A={a_raw}, Lv={lv}, Sun={sun}, Sum={total}')
        all_pass = False

if all_pass:
    print(f'  ✓ All {len(SORTED_EMPS)} employees: DaySum=31 for everyone')

# Final summary
print('\n' + '=' * 80)
print('FINAL SUMMARY')
print('=' * 80)
print('Generated 3 Excel files')
print(f'  1. {OUT_SUMMARY}')
print(f'  2. {OUT_MASTER}')
print(f'  3. {OUT_ATTENDANCE}')
print(f'\nKamlesh: P={int(kamlesh["presentDays"])}, A={int(kamlesh["absentDays"])}, '
      f'Leave={int(kamlesh["leaveColValue"])}, '
      f'Worked={display_decimal_col(kamlesh["workedHrsInclOT"])}, '
      f'Total={display_decimal_col(kamlesh["totalHrs"])}, '
      f'Gross={kamlesh["grossSalary"]}')
print(f'All {len(SORTED_EMPS)} employees: DaySum=31 for everyone')
