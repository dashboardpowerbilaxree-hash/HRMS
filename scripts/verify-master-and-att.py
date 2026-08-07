"""Verify Kamlesh's row in Payroll Master per-firm sheet + Attendance Tracker."""
import openpyxl

# ── File 2: Payroll_Master_July_2026.xlsx → per-firm sheet ──
print("[2] Payroll_Master_July_2026.xlsx → 'LAPL_July_2026_Sal' sheet")
wb = openpyxl.load_workbook('/home/z/my-project/download/Payroll_Master_July_2026.xlsx', data_only=False)
ws = wb['LAPL_July_2026_Sal']
# Print first 5 rows
for r in [1,2,3,4,5]:
    row_vals = [ws.cell(r, c).value for c in range(1, 16)]
    print(f"  R{r}: {row_vals}")
print()
# Find Kamlesh
for r in range(5, ws.max_row + 1):
    name = ws.cell(r, 2).value or ws.cell(r, 3).value
    if 'Kamlesh' in str(name):
        print(f"  Kamlesh @ row {r}:")
        for c in range(1, ws.max_column + 1):
            hdr = ws.cell(4, c).value or ws.cell(3, c).value or ws.cell(2, c).value or f"C{c}"
            v = ws.cell(r, c).value
            if v is not None or 'Present' in str(hdr) or 'Absent' in str(hdr) or 'Worked' in str(hdr) or 'Total' in str(hdr) or 'Sunday' in str(hdr) or 'Additional' in str(hdr):
                print(f"    {str(hdr)[:30]:30s}: {v}")
        break

# ── File 3: Attendance_Tracker_Monthly_July_2026.xlsx ──
print("\n[3] Attendance_Tracker_Monthly_July_2026.xlsx → 'LAPL_July_2026_Att' sheet")
wb = openpyxl.load_workbook('/home/z/my-project/download/Attendance_Tracker_Monthly_July_2026.xlsx', data_only=False)
ws = wb['LAPL_July_2026_Att']
# Print first 4 rows
for r in [1,2,3,4]:
    row_vals = [ws.cell(r, c).value for c in range(1, 12)]
    if any(v is not None for v in row_vals):
        print(f"  R{r}: {row_vals}")
print()
# Find Kamlesh
for r in range(5, ws.max_row + 1):
    name = ws.cell(r, 1).value or ws.cell(r, 2).value
    if 'Kamlesh' in str(name):
        print(f"  Kamlesh @ row {r}:")
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and v != '':
                h1 = ws.cell(1, c).value
                h2 = ws.cell(2, c).value
                h3 = ws.cell(3, c).value
                h4 = ws.cell(4, c).value
                print(f"    Col {c} ({h1}|{h2}|{h3}|{h4}): {v}")
        break
